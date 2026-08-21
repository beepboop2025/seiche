"""Official/open production adapters for monetary-area packs.

Each adapter is independently schedulable.  This module contains source
vocabulary and endpoint details; neither the market-neutral observation
contract nor the universal kernel imports it.
"""

from __future__ import annotations

import asyncio
import calendar as civil_calendar
import csv
import hashlib
import html
import io
import json
import logging
import os
import re
import urllib.parse
import zipfile
from datetime import UTC, date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from html.parser import HTMLParser
from typing import Any
from zoneinfo import ZoneInfo

import httpx

from seiche.markets.registry import MarketRegistry, default_registry
from seiche.repository import MarketRepository
from seiche.sources.base import SourcePolicyUnavailableError
from seiche.sources.canonical import (
    FetchedDocument,
    FunctionalCanonicalAdapter,
    ParsedPoint,
    get_documents,
)

USER_AGENT = "Seiche/0.9 (+https://seiche.info; official-data research collector)"

_HKMA_LIQUIDITY_URI = (
    "https://api.hkma.gov.hk/public/market-data-and-statistics/"
    "daily-monetary-statistics/daily-figures-interbank-liquidity"
)
_HKMA_RETRY_DELAYS_SECONDS = (1.0, 2.0)
_sleep = asyncio.sleep

_BOK_ECOS_BASE_URI = "https://ecos.bok.or.kr/api/StatisticSearch"
_BOK_ECOS_API_KEY_ENV = "SEICHE_BOK_ECOS_API_KEY"
_BOK_ECOS_PAGE_SIZE = 10_000
_BOK_ECOS_CREDENTIAL_URI_RE = re.compile(
    r"(https://ecos\.bok\.or\.kr/api/StatisticSearch/)[^/\s\"']+"
)
_BOK_ECOS_SERIES = {
    "KR.BOK.BASE_RATE": ("722Y001", "0101000"),
    "KR.BOK.CALL_OVERNIGHT_ALL": ("817Y002", "010101000"),
}


def _redact_bok_ecos_credential(value: object, api_key: str | None = None) -> str:
    """Remove an ECOS path credential before text crosses a diagnostic boundary."""

    text = str(value)
    key = api_key or os.getenv(_BOK_ECOS_API_KEY_ENV, "").strip()
    if key:
        text = text.replace(key, "REDACTED")
    return _BOK_ECOS_CREDENTIAL_URI_RE.sub(r"\1REDACTED", text)


class _BOKECOSLogRedactionFilter(logging.Filter):
    """Sanitize httpx's INFO request line before any handler can persist it."""

    def filter(self, record: logging.LogRecord) -> bool:
        message = record.getMessage()
        redacted = _redact_bok_ecos_credential(message)
        if record.exc_info is not None:
            rendered = logging.Formatter().formatException(record.exc_info)
            safe_rendered = _redact_bok_ecos_credential(rendered)
            if safe_rendered != rendered:
                redacted = f"{redacted}\n{safe_rendered}"
                record.exc_info = None
                record.exc_text = None
        if redacted != message:
            record.msg = redacted
            record.args = ()
        return True


_BOK_ECOS_LOG_REDACTION_FILTER = _BOKECOSLogRedactionFilter()
logging.getLogger("httpx").addFilter(_BOK_ECOS_LOG_REDACTION_FILTER)

_RBNZ_B2_XLSX_URI = (
    "https://www.rbnz.govt.nz/-/media/project/sites/rbnz/files/statistics/"
    "series/b/b2/hb2-daily-close.xlsx"
)
_RBNZ_B2_PAGE_URI = (
    "https://www.rbnz.govt.nz/en/statistics/series/exchange-and-interest-rates/"
    "wholesale-interest-rates"
)
_RBNZ_XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
_RBNZ_TERMS_URI = "https://www.rbnz.govt.nz/about-our-site/terms-of-use"
_RBNZ_ACCESS_APPROVAL_SHA256_ENV = "SEICHE_RBNZ_ACCESS_APPROVAL_SHA256"
_RBNZ_ACCESS_APPROVAL_VALID_UNTIL_ENV = "SEICHE_RBNZ_ACCESS_APPROVAL_VALID_UNTIL"
_RBNZ_ALLOWED_HOSTS = frozenset({"rbnz.govt.nz", "www.rbnz.govt.nz"})
_RBNZ_MAX_BODY_BYTES = 8 * 1024 * 1024
_RBNZ_MAX_XLSX_MEMBERS = 256
_RBNZ_MAX_XLSX_EXPANDED_BYTES = 64 * 1024 * 1024
_RBNZ_MAX_WORKBOOK_SHEETS = 16
_RBNZ_MAX_WORKBOOK_ROWS = 50_000
_RBNZ_MAX_WORKBOOK_COLUMNS = 256
_RBNZ_MAX_WORKBOOK_CELLS = 400_000
_RBNZ_MAX_APPROVAL_REVIEW_DAYS = 366
_RBNZ_TOTAL_RESPONSE_TIMEOUT_SECONDS = 90.0


class RBNZSourceUnavailableError(RuntimeError):
    """Both official RBNZ representations were unavailable or unusable."""


class RBNZAccessPolicyUnavailableError(
    RBNZSourceUnavailableError,
    SourcePolicyUnavailableError,
):
    """RBNZ access was withheld locally before any source request."""


class BOKECOSAccessPolicyUnavailableError(SourcePolicyUnavailableError):
    """BOK ECOS collection was withheld before any credential-bearing request."""


class BOKECOSSourceError(RuntimeError):
    """A credential-safe BOK ECOS source failure suitable for public faults."""


def _bok_ecos_api_key() -> str:
    """Return a non-sample ECOS key without ever placing it in diagnostics."""

    key = os.getenv(_BOK_ECOS_API_KEY_ENV, "").strip()
    if not key:
        raise BOKECOSAccessPolicyUnavailableError(
            f"BOK ECOS API key is required via {_BOK_ECOS_API_KEY_ENV}; "
            "no source request was made"
        )
    # ECOS' public ``sample`` key is deliberately capped at ten rows and can
    # silently turn a daily collection into an incomplete window. Production
    # collection therefore requires the individually issued alphanumeric key.
    if key.casefold() == "sample" or not re.fullmatch(r"[A-Za-z0-9]{8,128}", key):
        raise BOKECOSAccessPolicyUnavailableError(
            f"{_BOK_ECOS_API_KEY_ENV} must contain an issued BOK ECOS API key; "
            "the sample key is not a production credential"
        )
    return key


def _require_bok_ecos_api_key() -> None:
    _bok_ecos_api_key()


def bounded_date_windows(
    start: date,
    end: date,
    *,
    maximum_days: int,
) -> tuple[tuple[date, date], ...]:
    """Return non-overlapping inclusive windows within an upstream limit."""

    if maximum_days < 1:
        raise ValueError("maximum_days must be positive")
    if start > end:
        return ()
    windows = []
    cursor = start
    while cursor <= end:
        window_end = min(cursor + timedelta(days=maximum_days - 1), end)
        windows.append((cursor, window_end))
        cursor = window_end + timedelta(days=1)
    return tuple(windows)


def _number(value: object) -> Decimal | None:
    text = str(value).strip().replace(",", "")
    if text in {"", ".", "-", "NA", "N/A", "ND", "None", "nan"}:
        return None
    try:
        parsed = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    return parsed if parsed.is_finite() else None


def _date(value: object, *formats: str) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}(?:T.*)?", text):
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            pass
    return None


def _row_evidence(label: str, values: object) -> bytes:
    return json.dumps(
        {"label": label, "row": values},
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
        default=str,
    ).encode("utf-8")


def parse_fred_csv(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    instrument_id = document.label
    text = document.payload.decode("utf-8-sig")
    lines = text.splitlines()
    if len(lines) < 2:
        raise ValueError("FRED CSV contains no data rows")
    points: list[ParsedPoint] = []
    for raw_line in lines[1:]:
        row = next(csv.reader([raw_line]))
        if len(row) < 2:
            continue
        event_day = _date(row[0], "%Y-%m-%d")
        value = _number(row[1])
        if event_day is None or value is None:
            continue
        points.append(
            ParsedPoint(
                instrument_id,
                event_day,
                value,
                raw_line.encode("utf-8"),
            )
        )
    if not points:
        raise ValueError("FRED CSV contains no numeric observations")
    return tuple(points)


def parse_ecb_csv(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    text = document.payload.decode("utf-8-sig")
    lines = text.splitlines()
    if len(lines) < 2:
        raise ValueError("ECB SDMX CSV contains no data rows")
    header = next(csv.reader([lines[0]]))
    points: list[ParsedPoint] = []
    for raw_line in lines[1:]:
        values = next(csv.reader([raw_line]))
        if len(values) != len(header):
            continue
        row = dict(zip(header, values, strict=True))
        event_day = _date(row.get("TIME_PERIOD", ""), "%Y-%m-%d")
        value = _number(row.get("OBS_VALUE"))
        if event_day is None or value is None:
            continue
        revision = str(row.get("OBS_STATUS") or "").strip() or None
        points.append(
            ParsedPoint(
                document.label,
                event_day,
                value,
                raw_line.encode("utf-8"),
                revision_id=(
                    f"ecb:{revision}:{event_day.isoformat()}" if revision else None
                ),
            )
        )
    if not points:
        raise ValueError("ECB SDMX CSV contains no numeric observations")
    return tuple(points)


def parse_boe_csv(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    text = document.payload.decode("utf-8-sig")
    lines = [line for line in text.splitlines() if line.strip()]
    if len(lines) < 2:
        raise ValueError("Bank of England CSV contains no data rows")
    points: list[ParsedPoint] = []
    for raw_line in lines[1:]:
        row = next(csv.reader([raw_line]))
        if len(row) < 2:
            continue
        event_day = _date(row[0], "%d %b %Y", "%d/%b/%Y", "%Y-%m-%d")
        value = _number(row[1])
        if event_day is None or value is None:
            continue
        points.append(
            ParsedPoint(document.label, event_day, value, raw_line.encode("utf-8"))
        )
    if not points:
        raise ValueError("Bank of England CSV contains no numeric observations")
    return tuple(points)


def parse_boj_csv(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    text = document.payload.decode("utf-8-sig")
    lines = text.splitlines()
    code_row: list[str] | None = None
    update_day: date | None = None
    points: list[ParsedPoint] = []
    for raw_line in lines:
        row = next(csv.reader([raw_line]))
        if row and row[0] == "Series code":
            code_row = row[1:]
            continue
        if row and row[0] == "Last update":
            update_day = _date(row[1] if len(row) > 1 else "", "%Y/%m/%d")
            continue
        if not row:
            continue
        daily = _date(row[0], "%Y/%m/%d")
        monthly = None
        if daily is None and re.fullmatch(r"\d{4}/\d{2}", row[0].strip()):
            year, month = map(int, row[0].split("/"))
            monthly = date(year, month, civil_calendar.monthrange(year, month)[1])
        event_day = daily or monthly
        if event_day is None:
            continue
        if document.label == "JP.BOJ.CURRENT_ACCOUNTS":
            if not code_row:
                raise ValueError("BOJ monthly CSV has no series-code row")
            try:
                column = code_row.index("MD01'MABS1AN113") + 1
            except ValueError as exc:
                raise ValueError("BOJ current-account series column is absent") from exc
        else:
            column = 1
        if len(row) <= column:
            continue
        value = _number(row[column])
        if value is None:
            continue
        publication = None
        if (
            monthly is not None
            and update_day is not None
            and event_day.year == update_day.year
            and event_day.month == update_day.month - 1
        ):
            publication = datetime.combine(
                update_day,
                datetime.min.time(),
                tzinfo=ZoneInfo("Asia/Tokyo"),
            )
        points.append(
            ParsedPoint(
                document.label,
                event_day,
                value,
                raw_line.encode("utf-8"),
                source_publication_time=publication,
            )
        )
    if not points:
        raise ValueError("BOJ CSV contains no numeric observations")
    return tuple(points)


def _bok_ecos_payload_contains_credential(value: Any, credential: str) -> bool:
    if isinstance(value, str):
        return credential in value
    if isinstance(value, dict):
        return any(
            _bok_ecos_payload_contains_credential(key, credential)
            or _bok_ecos_payload_contains_credential(item, credential)
            for key, item in value.items()
        )
    if isinstance(value, list):
        return any(
            _bok_ecos_payload_contains_credential(item, credential) for item in value
        )
    return False


def _bok_ecos_search_payload(
    payload: bytes,
    *,
    reject_credential: str | None = None,
) -> tuple[dict[str, Any], list[Any]]:
    try:
        decoded = json.loads(payload)
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("BOK ECOS response is not valid JSON") from exc
    if not isinstance(decoded, dict):
        raise ValueError("BOK ECOS response is not a JSON object")
    if reject_credential and _bok_ecos_payload_contains_credential(
        decoded, reject_credential
    ):
        raise ValueError("BOK ECOS response echoed the API credential")
    result = decoded.get("RESULT")
    if isinstance(result, dict):
        code = str(result.get("CODE") or "unknown")
        message = str(result.get("MESSAGE") or "unspecified source error")
        raise ValueError(f"BOK ECOS returned {code}: {message}")
    search = decoded.get("StatisticSearch")
    if not isinstance(search, dict):
        raise ValueError("BOK ECOS response has no StatisticSearch object")
    rows = search.get("row")
    if not isinstance(rows, list):
        raise ValueError("BOK ECOS response has no row list")
    return search, rows


def parse_bok_ecos(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    """Parse only the two BOK-produced daily series declared by the KR pack."""

    try:
        expected_stat, expected_item = _BOK_ECOS_SERIES[document.label]
    except KeyError as exc:
        raise ValueError(f"unknown BOK ECOS document label {document.label!r}") from exc
    _, rows = _bok_ecos_search_payload(document.payload)
    points: list[ParsedPoint] = []
    for row in rows:
        if not isinstance(row, dict):
            raise ValueError("BOK ECOS row is not an object")
        if row.get("STAT_CODE") != expected_stat:
            raise ValueError(f"BOK ECOS row changed STAT_CODE for {document.label!r}")
        if row.get("ITEM_CODE1") != expected_item:
            raise ValueError(f"BOK ECOS row changed ITEM_CODE1 for {document.label!r}")
        if row.get("UNIT_NAME") != "Percent Per Annum":
            raise ValueError(f"BOK ECOS row changed UNIT_NAME for {document.label!r}")
        raw_time = str(row.get("TIME") or "").strip()
        event_day = (
            _date(raw_time, "%Y%m%d") if re.fullmatch(r"\d{8}", raw_time) else None
        )
        value = _number(row.get("DATA_VALUE"))
        if event_day is None or value is None:
            raise ValueError(f"BOK ECOS row is malformed for {document.label!r}")
        points.append(
            ParsedPoint(
                document.label,
                event_day,
                value,
                _row_evidence(document.label, row),
            )
        )
    if not points:
        raise ValueError("BOK ECOS response contains no numeric observations")
    return tuple(points)


def parse_cfets_csv(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    points: list[ParsedPoint] = []
    for raw_line in document.payload.decode("utf-8-sig").splitlines():
        row = next(csv.reader([raw_line]))
        if not row:
            continue
        event_day = _date(row[0], "%Y-%m-%d")
        if event_day is None:
            continue
        if document.label == "CN.CFETS.DR007":
            # CFETS' FDR007 fixing is the public depository-institution
            # seven-day repo benchmark corresponding to this pack role.
            column = 7
        else:
            column = 1
        if len(row) <= column:
            continue
        value = _number(row[column])
        if value is None:
            continue
        points.append(
            ParsedPoint(document.label, event_day, value, raw_line.encode("utf-8"))
        )
    if not points:
        raise ValueError("CFETS CSV contains no numeric observations")
    return tuple(points)


def parse_cfets_rates(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    if document.label == "CN.CFETS.DR007":
        return parse_cfets_csv(document)
    payload = json.loads(document.payload)
    records = payload.get("records")
    if not isinstance(records, list):
        raise ValueError("CFETS SHIBOR response has no records")
    points: list[ParsedPoint] = []
    for row in records:
        if not isinstance(row, dict):
            continue
        event_day = _date(row.get("showDateCN"), "%Y-%m-%d")
        value = _number(row.get("ON"))
        if event_day is None or value is None:
            continue
        points.append(
            ParsedPoint(
                "CN.CFETS.SHIBOR_ON",
                event_day,
                value,
                _row_evidence("SHIBOR_ON", row),
            )
        )
    if not points:
        raise ValueError("CFETS SHIBOR response contains no overnight values")
    return tuple(points)


def parse_nyfed_rates(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    payload = json.loads(document.payload)
    rows = payload.get("refRates")
    if not isinstance(rows, list):
        raise ValueError("NY Fed response has no refRates list")
    sofr_mapping = {
        # ``percentRate`` is the published transaction-weighted median.  P25
        # is a different distribution point and must never stand in for it.
        "percentRate": "US.NYFED.SOFR_MEDIAN",
        "percentPercentile99": "US.NYFED.SOFR_P99",
        "volumeInBillions": "US.NYFED.SOFR_VOLUME",
    }
    sofrai_mapping = {
        "average30day": "US.NYFED.SOFR_AVERAGE_30D",
        "average90day": "US.NYFED.SOFR_AVERAGE_90D",
        "average180day": "US.NYFED.SOFR_AVERAGE_180D",
        "index": "US.NYFED.SOFR_INDEX",
    }
    points: list[ParsedPoint] = []
    sofrai_seen: set[tuple[date, str]] = set()
    for row in rows:
        if not isinstance(row, dict):
            continue
        rate_type = str(row.get("type") or "").strip().upper()
        event_day = _date(row.get("effectiveDate"), "%Y-%m-%d")
        if event_day is None:
            continue
        revision = str(row.get("revisionIndicator") or "").strip()
        revision_token = re.sub(
            r"[^A-Za-z0-9._-]+", "-", revision or "unrevised"
        ).strip("-")
        if rate_type == "SOFR":
            mapping = sofr_mapping
            publication_time = None
            lineage_prefix = "nyfed"
        elif rate_type == "SOFRAI":
            mapping = sofrai_mapping
            # SOFR itself is labelled with the prior business day's effective
            # date, while the Averages and Index are labelled with their same-
            # day value date. They are published shortly after 08:00 ET on
            # that value date, so the adapter's T+1 SOFR clock cannot be used.
            publication_time = datetime.combine(
                event_day,
                time(8, 0),
                tzinfo=ZoneInfo("America/New_York"),
            ).astimezone(UTC)
            lineage_prefix = "nyfed:SOFRAI"
        else:
            continue
        for field, instrument in mapping.items():
            value = _number(row.get(field))
            if value is None:
                # Missing is unknown: never copy another horizon or emit zero.
                continue
            if rate_type == "SOFRAI":
                identity = (event_day, field)
                if identity in sofrai_seen:
                    raise ValueError(
                        "NY Fed response contains duplicate SOFRAI "
                        f"{field} for {event_day.isoformat()}"
                    )
                sofrai_seen.add(identity)
            row_evidence = _row_evidence(field, row)
            content_token = hashlib.sha256(row_evidence).hexdigest()[:16]
            points.append(
                ParsedPoint(
                    instrument,
                    event_day,
                    value,
                    row_evidence,
                    source_publication_time=publication_time,
                    # Bind the semantic source field and event explicitly even
                    # when upstream supplies no revision indicator. This keeps
                    # the median, tail, averaging horizons, and index distinct
                    # while retaining earlier values as ordinary revisions in
                    # the append-only canonical store.
                    revision_id=(
                        f"{lineage_prefix}:{field}:{event_day.isoformat()}:"
                        f"{revision_token or 'unrevised'}-{content_token}"
                    ),
                )
            )
    if not points:
        raise ValueError(
            "NY Fed response contains no SOFR distribution or averages/index rows"
        )
    return tuple(
        sorted(
            points,
            key=lambda point: (point.event_time, point.instrument_id),
        )
    )


def parse_nyfed_unsecured_rates(
    document: FetchedDocument,
) -> tuple[ParsedPoint, ...]:
    """Parse the canonical EFFR and OBFR distribution fields we can name.

    The NY Fed also publishes the 1st, 25th, and 75th percentiles and volumes.
    The canonical observation vocabulary does not yet have honest semantic
    roles for those fields, so they remain in the immutable raw capture rather
    than being mislabeled or coerced into a repo-volume series.
    """

    payload = json.loads(document.payload)
    rows = payload.get("refRates")
    if not isinstance(rows, list):
        raise ValueError("NY Fed unsecured response has no refRates list")
    mappings = {
        "EFFR": {
            "percentRate": "US.NYFED.EFFR_MEDIAN",
            "percentPercentile99": "US.NYFED.EFFR_P99",
        },
        "OBFR": {
            "percentRate": "US.NYFED.OBFR_MEDIAN",
            "percentPercentile99": "US.NYFED.OBFR_P99",
        },
    }
    points: list[ParsedPoint] = []
    seen: set[tuple[str, date, str]] = set()
    for row in rows:
        if not isinstance(row, dict):
            continue
        rate_type = str(row.get("type") or "").strip().upper()
        mapping = mappings.get(rate_type)
        if mapping is None:
            continue
        event_day = _date(row.get("effectiveDate"), "%Y-%m-%d")
        if event_day is None:
            continue
        revision = str(row.get("revisionIndicator") or "").strip()
        revision_token = re.sub(
            r"[^A-Za-z0-9._-]+", "-", revision or "unrevised"
        ).strip("-")
        for field, instrument in mapping.items():
            value = _number(row.get(field))
            if value is None:
                # An absent distribution statistic is unknown, never zero and
                # never copied from another percentile.
                continue
            identity = (rate_type, event_day, field)
            if identity in seen:
                raise ValueError(
                    "NY Fed unsecured response contains duplicate "
                    f"{rate_type} {field} for {event_day.isoformat()}"
                )
            seen.add(identity)
            row_evidence = _row_evidence(f"{rate_type}.{field}", row)
            content_token = hashlib.sha256(row_evidence).hexdigest()[:16]
            points.append(
                ParsedPoint(
                    instrument,
                    event_day,
                    value,
                    row_evidence,
                    revision_id=(
                        f"nyfed:{rate_type}:{field}:{event_day.isoformat()}:"
                        f"{revision_token or 'unrevised'}-{content_token}"
                    ),
                )
            )
    if not points:
        raise ValueError(
            "NY Fed unsecured response contains no canonical EFFR or OBFR "
            "distribution rows"
        )
    return tuple(
        sorted(
            points,
            key=lambda point: (point.event_time, point.instrument_id),
        )
    )


def parse_nyfed_facilities(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    payload = json.loads(document.payload)
    operations = (payload.get("repo") or {}).get("operations")
    if not isinstance(operations, list):
        raise ValueError("NY Fed response has no repo operations")
    grouped: dict[date, list[dict[str, Any]]] = {}
    for operation in operations:
        if not isinstance(operation, dict):
            continue
        event_day = _date(operation.get("operationDate"), "%Y-%m-%d")
        # Missing is not zero. Explicit numeric zero remains a valid no-takeup
        # observation because the official result reported it.
        if event_day is None or operation.get("totalAmtAccepted") is None:
            continue
        grouped.setdefault(event_day, []).append(operation)
    points = []
    for event_day, rows in grouped.items():
        amounts = [_number(row["totalAmtAccepted"]) for row in rows]
        if any(value is None for value in amounts):
            continue
        # API amounts are dollars; pack normalization expects USD billions.
        value = sum(amounts, Decimal(0)) / Decimal(1_000_000_000)  # type: ignore[arg-type]
        points.append(
            ParsedPoint(
                "US.NYFED.SRF_TAKEUP",
                event_day,
                value,
                _row_evidence("totalAmtAccepted", rows),
            )
        )
    if not points:
        raise ValueError("NY Fed response contains no explicit facility results")
    return tuple(points)


def parse_fiscal_tga(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    payload = json.loads(document.payload)
    rows = payload.get("data")
    if not isinstance(rows, list):
        raise ValueError("FiscalData response has no data list")
    priority = {
        "Treasury General Account (TGA) Opening Balance": 0,
        "Treasury General Account (TGA)": 1,
        "Federal Reserve Account": 2,
    }
    selected: dict[date, tuple[int, dict[str, Any]]] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        event_day = _date(row.get("record_date"), "%Y-%m-%d")
        value = _number(row.get("open_today_bal"))
        account = str(row.get("account_type") or "")
        if event_day is None or value is None or account not in priority:
            continue
        candidate = (priority[account], row)
        if event_day not in selected or candidate[0] < selected[event_day][0]:
            selected[event_day] = candidate
    points = []
    for event_day, (_, row) in sorted(selected.items()):
        # FiscalData reports USD millions; pack source unit is USD billions.
        value = _number(row["open_today_bal"])
        points.append(
            ParsedPoint(
                "US.TREASURY.GOVERNMENT_CASH",
                event_day,
                value / Decimal(1000),  # type: ignore[operator]
                _row_evidence("tga", row),
            )
        )
    if not points:
        raise ValueError("FiscalData response contains no TGA rows")
    return tuple(points)


def _workbook(document: FetchedDocument):
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - deployment extra
        raise RuntimeError(
            "official workbook adapter needs seiche[collectors]"
        ) from exc
    return openpyxl.load_workbook(
        io.BytesIO(document.payload),
        read_only=True,
        data_only=True,
    )


def _workbook_series_rows(document: FetchedDocument) -> tuple[list[str], list[tuple]]:
    workbook = _workbook(document)
    worksheet = (
        workbook["Data"] if "Data" in workbook.sheetnames else workbook.worksheets[0]
    )
    series_ids: list[str] | None = None
    rows: list[tuple] = []
    for values in worksheet.iter_rows(values_only=True):
        row = tuple(values)
        if row and str(row[0]).strip().lower() == "series id":
            series_ids = [
                str(value).strip() if value is not None else "" for value in row
            ]
            continue
        if series_ids is not None and row and isinstance(row[0], (date, datetime)):
            rows.append(row)
    if series_ids is None or not rows:
        raise ValueError("official workbook has no Series ID/data rows")
    return series_ids, rows


def parse_rba_workbook(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    series_ids, rows = _workbook_series_rows(document)
    mappings = {
        "rba_cash": {"FIRMMCRID": "AU.RBA.AONIA"},
        "rba_policy_daily": {"FIRMMCRTD": "AU.RBA.CASH_TARGET"},
        "rba_policy_changes": {
            "ARBAMPNRPESB": "AU.RBA.ES_FLOOR",
            "ARBAMPNORR": "AU.RBA.STANDING_CEILING",
        },
    }
    try:
        selected = mappings[document.label]
    except KeyError as exc:
        raise ValueError(f"unknown RBA document label {document.label!r}") from exc
    columns = {
        index: selected[series_id]
        for index, series_id in enumerate(series_ids)
        if series_id in selected
    }
    if not columns:
        raise ValueError("RBA workbook no longer contains expected series IDs")
    points: list[ParsedPoint] = []
    for row in rows:
        event_day = _date(row[0])
        if event_day is None:
            continue
        for column, instrument in columns.items():
            if len(row) <= column:
                continue
            value = _number(row[column])
            if value is None:
                continue
            evidence = _row_evidence(
                series_ids[column],
                [event_day.isoformat(), str(row[column])],
            )
            points.append(ParsedPoint(instrument, event_day, value, evidence))
    if not points:
        raise ValueError("RBA workbook contains no expected numeric observations")
    return tuple(points)


class _TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[list[str]] = []
        self.tables: list[list[list[str]]] = []
        self._table: list[list[str]] | None = None
        self._table_depth = 0
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag.lower() == "table":
            if self._table_depth == 0:
                self._table = []
            self._table_depth += 1
        elif tag.lower() == "tr":
            self._row = []
        elif tag.lower() in {"td", "th"} and self._row is not None:
            self._cell = []

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"td", "th"} and self._cell is not None:
            value = " ".join("".join(self._cell).replace("\xa0", " ").split())
            if self._row is not None:
                self._row.append(value)
            self._cell = None
        elif tag.lower() == "tr" and self._row is not None:
            self.rows.append(self._row)
            if self._table is not None:
                self._table.append(self._row)
            self._row = None
        elif tag.lower() == "table" and self._table_depth:
            self._table_depth -= 1
            if self._table_depth == 0 and self._table is not None:
                self.tables.append(self._table)
                self._table = None


class _HiddenInputParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.values: dict[str, str] = {}

    def handle_starttag(self, tag: str, attrs) -> None:
        values = dict(attrs)
        if (
            tag.lower() == "input"
            and values.get("type", "").lower() == "hidden"
            and values.get("name")
        ):
            self.values[values["name"]] = values.get("value", "")


def _table_rows(payload: bytes) -> list[list[str]]:
    parser = _TableParser()
    parser.feed(payload.decode("utf-8-sig", errors="replace"))
    return parser.rows


def _html_tables(payload: bytes) -> list[list[list[str]]]:
    parser = _TableParser()
    parser.feed(payload.decode("utf-8-sig", errors="replace"))
    return parser.tables


def parse_mas_sora(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    year: int | None = None
    month: int | None = None
    points: list[ParsedPoint] = []
    for row in _table_rows(document.payload):
        if len(row) != 6:
            continue
        if row[0].isdigit():
            year = int(row[0])
        if row[1]:
            try:
                month = datetime.strptime(row[1][:3], "%b").month
            except ValueError:
                continue
        if year is None or month is None or not row[2].isdigit():
            continue
        event_day = date(year, month, int(row[2]))
        publication_day = _date(row[3], "%d %b %Y")
        value = _number(row[4])
        if publication_day is None or value is None:
            continue
        publication = datetime.combine(
            publication_day,
            datetime.min.time().replace(hour=9),
            tzinfo=ZoneInfo("Asia/Singapore"),
        )
        points.append(
            ParsedPoint(
                "SG.MAS.SORA",
                event_day,
                value,
                _row_evidence("sora", row),
                source_publication_time=publication,
            )
        )
    if not points:
        raise ValueError("MAS result contains no SORA rows")
    return tuple(points)


def parse_mas_rates(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    year: int | None = None
    month: int | None = None
    points: list[ParsedPoint] = []
    for row in _table_rows(document.payload):
        if len(row) != 5:
            continue
        if row[0].isdigit():
            year = int(row[0])
        if row[1]:
            try:
                month = datetime.strptime(row[1][:3], "%b").month
            except ValueError:
                continue
        if year is None or month is None or not row[2].isdigit():
            continue
        event_day = date(year, month, int(row[2]))
        for column, instrument in (
            (3, "SG.MAS.STANDING_DEPOSIT"),
            (4, "SG.MAS.STANDING_BORROWING"),
        ):
            value = _number(row[column])
            if value is not None:
                points.append(
                    ParsedPoint(
                        instrument,
                        event_day,
                        value,
                        _row_evidence(instrument, row),
                    )
                )
    if not points:
        raise ValueError("MAS result contains no standing-facility rows")
    return tuple(points)


def parse_hkma_json(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    payload = json.loads(document.payload)
    records = (payload.get("result") or {}).get("records")
    if not isinstance(records, list):
        raise ValueError("HKMA response has no records")
    mapping = {
        "disc_win_base_rate": "HK.HKMA.BASE_RATE",
        "closing_balance": "HK.HKMA.AGGREGATE_BALANCE",
    }
    points: list[ParsedPoint] = []
    for row in records:
        if not isinstance(row, dict):
            continue
        event_day = _date(row.get("end_of_date"), "%Y-%m-%d")
        if event_day is None:
            continue
        for field, instrument in mapping.items():
            value = _number(row.get(field))
            if value is None:
                continue
            points.append(
                ParsedPoint(
                    instrument,
                    event_day,
                    value,
                    _row_evidence(field, row),
                )
            )
    if not points:
        raise ValueError("HKMA response contains no mapped rows")
    return tuple(points)


def _html_text(fragment: str) -> str:
    return " ".join(
        html.unescape(re.sub(r"<[^>]+>", " ", fragment)).replace("\xa0", " ").split()
    )


def _html_row_by_id(text: str, row_id: str) -> tuple[list[str], bytes] | None:
    for match in re.finditer(r"<tr\b[^>]*>.*?</tr>", text, re.IGNORECASE | re.DOTALL):
        block = match.group(0)
        if not re.search(rf'id=["\']{re.escape(row_id)}["\']', block, re.IGNORECASE):
            continue
        cells = [
            _html_text(value)
            for value in re.findall(
                r"<(?:td|th)\b[^>]*>(.*?)</(?:td|th)>",
                block,
                re.IGNORECASE | re.DOTALL,
            )
        ]
        return cells, block.encode("utf-8")
    return None


def parse_rbi_html(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    text = document.payload.decode("utf-8-sig", errors="replace")
    points: list[ParsedPoint] = []
    if document.label == "rbi_home":
        event_match = re.search(
            r"(?:As at|as on)\s+(?:\d{1,2}[.:]\d{2}\s*(?:am|pm)\s+of\s+)?"
            r"([A-Z][a-z]+\s+\d{1,2},\s+\d{4})",
            _html_text(text),
        )
        event_day = _date(event_match.group(1), "%B %d, %Y") if event_match else None
        if event_day is None:
            return ()
        plain = _html_text(text)
        for label, instrument in (
            ("Policy Repo Rate", "IN.RBI.POLICY_REPO"),
            ("91 day T-bills", "IN.RBI.TBILL_3M"),
        ):
            match = re.search(
                rf"{re.escape(label)}\s*(?::)?\s*([0-9]+(?:\.[0-9]+)?)\s*%?",
                plain,
                re.IGNORECASE,
            )
            if match:
                points.append(
                    ParsedPoint(
                        instrument,
                        event_day,
                        match.group(1),
                        _row_evidence(label, match.group(0)),
                    )
                )
        return tuple(points)

    event_match = re.search(r"Money Market Operations as on\s+([^<]+)</b>", text, re.I)
    event_day = (
        _date(_html_text(event_match.group(1)), "%B %d, %Y") if event_match else None
    )
    if event_day is None:
        raise ValueError("RBI page has no money-market event date")

    def add_from_row(row_id: str, column: int, instrument: str) -> None:
        found = _html_row_by_id(text, row_id)
        if found is None:
            return
        cells, evidence = found
        if column >= len(cells):
            return
        value = _number(cells[column])
        if value is not None:
            points.append(ParsedPoint(instrument, event_day, value, evidence))

    add_from_row("OSCallMoney", 2, "IN.MARKET.CALL_WAR")
    add_from_row("OSTriparty", 1, "IN.RBI.TRIPARTY_REPO_VOLUME")
    add_from_row("MSF3", 5, "IN.RBI.MSF")
    add_from_row("SDF2", 5, "IN.RBI.SDF")
    add_from_row("Netliquidityinjectedoutstandingtoday", 4, "IN.RBI.SYSTEM_LIQUIDITY")
    add_from_row("CashBalRBI", 2, "IN.RBI.CASH_BALANCES")
    add_from_row("GovernmentIndiaSurplusCashBalance", 2, "IN.GOVERNMENT.CASH_BALANCE")

    facility_rows = [
        found
        for row_id in ("MSF3", "SDF2")
        if (found := _html_row_by_id(text, row_id)) is not None
    ]
    amounts = [_number(cells[4]) for cells, _ in facility_rows if len(cells) > 4]
    if amounts and all(value is not None for value in amounts):
        points.append(
            ParsedPoint(
                "IN.RBI.FACILITY_TAKEUP",
                event_day,
                sum(amounts, Decimal(0)),  # type: ignore[arg-type]
                b"\n".join(evidence for _, evidence in facility_rows),
            )
        )
    if not points:
        raise ValueError("RBI page contains no mapped money-market rows")
    return tuple(points)


_RBNZ_COLUMN_HEADINGS = {
    "NZ.RBNZ.OCR": "official cash rate",
    "NZ.RBNZ.OVERNIGHT_DEPOSIT": "overnight deposit",
    "NZ.RBNZ.OVERNIGHT_REVERSE_REPO": "overnight reverse",
}
_RBNZ_SERIES_IDS = {
    "INM.DP1.N": "NZ.RBNZ.OCR",
    "INM.DD1.N": "NZ.RBNZ.OVERNIGHT_DEPOSIT",
    "INM.DD2.N": "NZ.RBNZ.OVERNIGHT_REVERSE_REPO",
}


def _rbnz_html_data_rows(payload: bytes) -> tuple[list[list[str]], dict[int, str]]:
    """Locate the declared B2 table instead of accepting any dated HTML table."""

    for rows in _html_tables(payload):
        for header_index, row in enumerate(rows):
            normalized = [" ".join(value.lower().split()) for value in row]
            if not normalized or normalized[0] != "date":
                continue
            columns: dict[int, str] = {}
            for column, heading in enumerate(normalized):
                for instrument, needle in _RBNZ_COLUMN_HEADINGS.items():
                    if needle in heading:
                        columns[column] = instrument
            if len(columns) == len(_RBNZ_COLUMN_HEADINGS) and set(
                columns.values()
            ) == set(_RBNZ_COLUMN_HEADINGS):
                return rows[header_index + 1 :], columns
    raise ValueError("RBNZ HTML page has no expected B2 interest-rate table header")


def _validate_rbnz_workbook_archive(payload: bytes) -> None:
    """Bound ZIP expansion before openpyxl reads an upstream workbook."""

    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            members = archive.infolist()
    except (OSError, zipfile.BadZipFile) as exc:
        raise ValueError("RBNZ workbook is not a valid XLSX ZIP archive") from exc
    if not members or len(members) > _RBNZ_MAX_XLSX_MEMBERS:
        raise ValueError("RBNZ workbook exceeds the XLSX member limit")
    if len({member.filename for member in members}) != len(members):
        raise ValueError("RBNZ workbook contains duplicate XLSX member names")
    expanded_bytes = 0
    for member in members:
        parts = member.filename.replace("\\", "/").split("/")
        if member.filename.startswith("/") or ".." in parts:
            raise ValueError("RBNZ workbook contains an unsafe XLSX member name")
        if member.flag_bits & 0x1:
            raise ValueError("RBNZ workbook contains an encrypted XLSX member")
        if member.compress_type not in {zipfile.ZIP_STORED, zipfile.ZIP_DEFLATED}:
            raise ValueError("RBNZ workbook uses an unsupported ZIP compression method")
        expanded_bytes += member.file_size
        if expanded_bytes > _RBNZ_MAX_XLSX_EXPANDED_BYTES:
            raise ValueError("RBNZ workbook exceeds the XLSX expansion limit")


def _bounded_rbnz_worksheets(
    document: FetchedDocument,
) -> list[tuple[str, list[tuple]]]:
    _validate_rbnz_workbook_archive(document.payload)
    workbook = _workbook(document)
    try:
        if len(workbook.worksheets) > _RBNZ_MAX_WORKBOOK_SHEETS:
            raise ValueError("RBNZ workbook exceeds the worksheet limit")
        worksheets: list[tuple[str, list[tuple]]] = []
        row_count = 0
        cell_count = 0
        for sheet in workbook.worksheets:
            if (
                sheet.max_column is not None
                and sheet.max_column > _RBNZ_MAX_WORKBOOK_COLUMNS
            ):
                raise ValueError("RBNZ workbook exceeds the column limit")
            if (
                sheet.max_row is not None
                and row_count + sheet.max_row > _RBNZ_MAX_WORKBOOK_ROWS
            ):
                raise ValueError("RBNZ workbook exceeds the row limit")
            rows: list[tuple] = []
            for values in sheet.iter_rows(values_only=True):
                row = tuple(values)
                row_count += 1
                cell_count += len(row)
                if cell_count > _RBNZ_MAX_WORKBOOK_CELLS:
                    raise ValueError("RBNZ workbook exceeds the cell limit")
                if row_count > _RBNZ_MAX_WORKBOOK_ROWS:
                    raise ValueError("RBNZ workbook exceeds the row limit")
                if len(row) > _RBNZ_MAX_WORKBOOK_COLUMNS:
                    raise ValueError("RBNZ workbook exceeds the column limit")
                rows.append(row)
            worksheets.append((sheet.title, rows))
        return worksheets
    finally:
        workbook.close()


def _rbnz_workbook_data_rows(
    document: FetchedDocument,
) -> tuple[list[tuple], dict[int, str]]:
    """Locate one unambiguous B2 worksheet using IDs, then legacy headings."""

    worksheets = _bounded_rbnz_worksheets(document)
    expected = set(_RBNZ_COLUMN_HEADINGS)
    candidates: list[tuple[str, list[tuple], dict[int, str]]] = []
    saw_series_layout = any(
        " ".join(title.lower().split()) in {"series definitions", "table description"}
        for title, _ in worksheets
    )

    for title, rows in worksheets:
        for index, row in enumerate(rows):
            marker = " ".join(str(row[0] or "").lower().split()) if row else ""
            values = {str(value).strip() for value in row}
            if marker.startswith("series id") or values.intersection(_RBNZ_SERIES_IDS):
                saw_series_layout = True
            if marker != "series id":
                continue
            columns = {
                column: _RBNZ_SERIES_IDS[str(value).strip()]
                for column, value in enumerate(row)
                if str(value).strip() in _RBNZ_SERIES_IDS
            }
            if len(columns) != len(expected) or set(columns.values()) != expected:
                continue
            data_rows = [
                tuple(candidate)
                for candidate in rows[index + 1 :]
                if candidate
                and _date(candidate[0]) is not None
                and any(
                    len(candidate) > column and _number(candidate[column]) is not None
                    for column in columns
                )
            ]
            if data_rows:
                candidates.append((title, data_rows, columns))

    # A modern workbook declaring Series IDs must satisfy the exact ID contract;
    # never downgrade it to fuzzy display-heading matching.
    if not saw_series_layout:
        for title, rows in worksheets:
            for index, row in enumerate(rows):
                if not row or str(row[0]).strip().lower() != "date":
                    continue
                if any(
                    any(value is not None and str(value).strip() for value in prior)
                    for prior in rows[:index]
                ):
                    continue
                following = next(
                    (
                        candidate
                        for candidate in rows[index + 1 :]
                        if candidate
                        and any(
                            value is not None and str(value).strip()
                            for value in candidate
                        )
                    ),
                    (),
                )
                if not following or _date(following[0]) is None:
                    continue
                columns: dict[int, str] = {}
                for column, heading in enumerate(row):
                    normalized = " ".join(str(heading or "").lower().split())
                    for instrument, needle in _RBNZ_COLUMN_HEADINGS.items():
                        if needle in normalized:
                            columns[column] = instrument
                if len(columns) != len(expected) or set(columns.values()) != expected:
                    continue
                data_rows = [
                    tuple(candidate)
                    for candidate in rows[index + 1 :]
                    if candidate
                    and _date(candidate[0]) is not None
                    and any(
                        len(candidate) > column
                        and _number(candidate[column]) is not None
                        for column in columns
                    )
                ]
                if data_rows:
                    candidates.append((title, data_rows, columns))

    if len(candidates) != 1:
        titles = ", ".join(title for title, _, _ in candidates) or "none"
        raise ValueError(
            "RBNZ workbook must contain exactly one usable B2 data sheet; "
            f"found {len(candidates)} ({titles})"
        )
    _, rows, columns = candidates[0]
    return rows, columns


def parse_rbnz(document: FetchedDocument) -> tuple[ParsedPoint, ...]:
    points: list[ParsedPoint] = []
    if (
        "spreadsheet" in document.media_type
        or document.source_uri.endswith(".xlsx")
        or document.payload.startswith(b"PK\x03\x04")
    ):
        rows, columns = _rbnz_workbook_data_rows(document)
        for row in rows:
            if not row:
                continue
            event_day = _date(row[0])
            if event_day is None:
                continue
            for column, instrument in columns.items():
                if len(row) <= column or (value := _number(row[column])) is None:
                    continue
                points.append(
                    ParsedPoint(
                        instrument,
                        event_day,
                        value,
                        _row_evidence(instrument, [event_day, row[column]]),
                    )
                )
    else:
        rows, columns = _rbnz_html_data_rows(document.payload)
        for row in rows:
            if not row:
                continue
            event_day = _date(row[0], "%d %b %Y", "%d %B %Y")
            if event_day is None:
                continue
            for column, instrument in columns.items():
                if len(row) <= column:
                    continue
                value = _number(row[column])
                if value is not None:
                    points.append(
                        ParsedPoint(
                            instrument,
                            event_day,
                            value,
                            _row_evidence(
                                instrument,
                                [event_day.isoformat(), str(row[column])],
                            ),
                        )
                    )
    allowed = {
        "rbnz_policy": {"NZ.RBNZ.OCR"},
        "rbnz_wholesale": {
            "NZ.RBNZ.OVERNIGHT_DEPOSIT",
            "NZ.RBNZ.OVERNIGHT_REVERSE_REPO",
        },
    }[document.label]
    selected = tuple(point for point in points if point.instrument_id in allowed)
    if not selected:
        raise ValueError("RBNZ source contains no mapped rows")
    return selected


async def _mas_documents(
    client: httpx.AsyncClient,
    *,
    label: str,
    columns: tuple[int, ...],
    start_year: int,
    end_year: int,
    end_month: int,
) -> tuple[FetchedDocument, ...]:
    uri = "https://eservices.mas.gov.sg/statistics/dir/DomesticInterestRates.aspx"
    response = await client.get(uri)
    response.raise_for_status()
    hidden = _HiddenInputParser()
    hidden.feed(response.text)
    form = {
        **hidden.values,
        "ctl00$ContentPlaceHolder1$StartYearDropDownList": str(start_year),
        "ctl00$ContentPlaceHolder1$EndYearDropDownList": str(end_year),
        "ctl00$ContentPlaceHolder1$StartMonthDropDownList": "1",
        "ctl00$ContentPlaceHolder1$EndMonthDropDownList": str(end_month),
        "ctl00$ContentPlaceHolder1$Button1": "Display",
    }
    for column in columns:
        form[f"ctl00$ContentPlaceHolder1$ColumnsCheckBoxList${column}"] = "on"
    result = await client.post(uri, data=form)
    result.raise_for_status()
    if "ResultsContainerPanel" not in result.text:
        raise ValueError("MAS form returned no result table")
    return (
        FetchedDocument(
            str(result.url),
            "text/html",
            result.content,
            label,
        ),
    )


def _response_media_type(
    response: httpx.Response,
    default: str = "application/octet-stream",
) -> str:
    return response.headers.get("content-type", default).split(";", 1)[0].lower()


async def _fetch_hkma_documents(
    client: httpx.AsyncClient,
) -> tuple[FetchedDocument, ...]:
    """Fetch HKMA liquidity within one transport/5xx retry budget."""

    for attempt in range(len(_HKMA_RETRY_DELAYS_SECONDS) + 1):
        try:
            response = await client.get(
                _HKMA_LIQUIDITY_URI,
                params={"pagesize": 1000},
                headers={"Accept": "application/json", "User-Agent": USER_AGENT},
            )
        except httpx.RequestError:
            if attempt == len(_HKMA_RETRY_DELAYS_SECONDS):
                raise
        else:
            if not 500 <= response.status_code < 600:
                response.raise_for_status()
                return (
                    FetchedDocument(
                        str(response.url),
                        _response_media_type(response),
                        response.content,
                        "hkma_liquidity",
                    ),
                )
            if attempt == len(_HKMA_RETRY_DELAYS_SECONDS):
                response.raise_for_status()
        await _sleep(_HKMA_RETRY_DELAYS_SECONDS[attempt])
    raise RuntimeError(
        "HKMA request loop exhausted without a response"
    )  # pragma: no cover


def _rbnz_failure_detail(exc: Exception) -> str:
    if isinstance(exc, httpx.HTTPStatusError):
        return f"HTTP {exc.response.status_code} from {exc.response.url}"
    if isinstance(exc, httpx.RequestError):
        return f"{type(exc).__name__} from {exc.request.url}: {exc}"
    return f"{type(exc).__name__}: {exc}"


def _rbnz_access_today() -> date:
    return datetime.now(UTC).date()


def _require_rbnz_access_approval() -> None:
    """Require a bounded operator attestation to RBNZ's written permission."""

    approval_hash = os.getenv(_RBNZ_ACCESS_APPROVAL_SHA256_ENV, "").strip().lower()
    valid_until_raw = os.getenv(_RBNZ_ACCESS_APPROVAL_VALID_UNTIL_ENV, "").strip()
    try:
        valid_until = date.fromisoformat(valid_until_raw)
    except ValueError:
        valid_until = None
    if not re.fullmatch(r"[0-9a-f]{64}", approval_hash) or valid_until is None:
        raise RBNZAccessPolicyUnavailableError(
            "RBNZ automated access is disabled before any request: prior written "
            f"permission is required by {_RBNZ_TERMS_URI}; configure an approval "
            "artifact SHA-256 and ISO valid-until date only after approval"
        )
    today = _rbnz_access_today()
    review_days = (valid_until - today).days
    if review_days < 0:
        raise RBNZAccessPolicyUnavailableError(
            "RBNZ automated-access approval review has expired; no request was made"
        )
    if review_days > _RBNZ_MAX_APPROVAL_REVIEW_DAYS:
        raise RBNZAccessPolicyUnavailableError(
            "RBNZ automated-access approval must be reviewed within 366 days; "
            "no request was made"
        )


def _rbnz_request_headers(*, navigation: bool) -> dict[str, str]:
    """Identify Seiche honestly on an operator-approved RBNZ connection."""

    headers = {
        "Accept-Encoding": "identity",
        "Accept-Language": "en-NZ,en;q=0.8",
        "User-Agent": USER_AGENT,
    }
    if navigation:
        headers["Accept"] = "text/html, application/xhtml+xml;q=0.9"
    else:
        headers.update(
            {
                "Accept": f"{_RBNZ_XLSX_MEDIA_TYPE}, application/octet-stream;q=0.9",
                "Referer": _RBNZ_B2_PAGE_URI,
            }
        )
    return headers


def _validate_rbnz_url(url: str) -> None:
    """Accept only the two canonical RBNZ HTTPS hostnames and default port."""

    try:
        parsed = urllib.parse.urlsplit(url)
        port = parsed.port
    except ValueError as exc:
        raise ValueError(f"invalid RBNZ response URL: {url!r}") from exc
    host = (parsed.hostname or "").lower().rstrip(".")
    if (
        parsed.scheme != "https"
        or host not in _RBNZ_ALLOWED_HOSTS
        or port not in {None, 443}
        or parsed.username is not None
        or parsed.password is not None
    ):
        raise ValueError(f"RBNZ response left its official HTTPS origin: {url}")


async def _read_rbnz_url(
    client: httpx.AsyncClient,
    uri: str,
    headers: dict[str, str],
) -> tuple[str, str, bytes]:
    """Perform one bounded request without following a redirect."""

    _validate_rbnz_url(uri)
    try:
        async with asyncio.timeout(_RBNZ_TOTAL_RESPONSE_TIMEOUT_SECONDS):
            async with client.stream(
                "GET", uri, headers=headers, follow_redirects=False
            ) as response:
                final_uri = str(response.url)
                _validate_rbnz_url(final_uri)
                if response.status_code != 200:
                    response.raise_for_status()
                    raise ValueError(f"HTTP {response.status_code} from {final_uri}")
                content_encoding = (
                    response.headers.get("content-encoding", "").strip().lower()
                )
                if content_encoding not in {"", "identity"}:
                    raise ValueError(
                        "RBNZ response used an unsupported transport content encoding"
                    )
                length = response.headers.get("content-length")
                if length is not None:
                    declared = int(length)
                    if declared < 0 or declared > _RBNZ_MAX_BODY_BYTES:
                        raise ValueError("RBNZ response exceeds the 8 MiB body limit")
                payload = bytearray()
                async for chunk in response.aiter_raw():
                    payload.extend(chunk)
                    if len(payload) > _RBNZ_MAX_BODY_BYTES:
                        raise ValueError("RBNZ response exceeds the 8 MiB body limit")
                return final_uri, _response_media_type(response), bytes(payload)
    except TimeoutError as exc:
        raise ValueError("RBNZ response exceeded the total response deadline") from exc


def _validate_rbnz_document(document: FetchedDocument) -> None:
    """Prove a candidate can emit the adapter's declared B2 instruments."""

    try:
        parse_rbnz(document)
    except Exception as exc:  # noqa: BLE001 - normalize source/parser failures
        raise ValueError(
            f"{document.source_uri} has no usable {document.label} B2 data: "
            f"{type(exc).__name__}: {exc}"
        ) from exc


async def _fetch_rbnz_document(
    client: httpx.AsyncClient,
    label: str,
    *,
    historical_backfill: bool,
) -> tuple[FetchedDocument, ...]:
    """Try the official B2 workbook, then its canonical official HTML table."""

    _require_rbnz_access_approval()
    try:
        source_uri, media_type, payload = await _read_rbnz_url(
            client,
            _RBNZ_B2_XLSX_URI,
            _rbnz_request_headers(navigation=False),
        )
        _validate_rbnz_url(source_uri)
        if not payload.startswith(b"PK\x03\x04"):
            raise ValueError(
                f"{source_uri} returned {media_type} instead of an XLSX workbook"
            )
        document = FetchedDocument(
            source_uri,
            media_type or _RBNZ_XLSX_MEDIA_TYPE,
            payload,
            label,
        )
        _validate_rbnz_document(document)
    except (httpx.HTTPError, ValueError) as exc:
        primary_failure = _rbnz_failure_detail(exc)
    else:
        return (document,)

    try:
        source_uri, media_type, payload = await _read_rbnz_url(
            client,
            _RBNZ_B2_PAGE_URI,
            _rbnz_request_headers(navigation=True),
        )
        _validate_rbnz_url(source_uri)
        document = FetchedDocument(
            source_uri,
            media_type or "text/html",
            payload,
            label,
        )
        _validate_rbnz_document(document)
    except (httpx.HTTPError, ValueError) as exc:
        fallback_failure = _rbnz_failure_detail(exc)
        raise RBNZSourceUnavailableError(
            f"RBNZ B2 workbook failed ({primary_failure}); canonical HTML "
            f"fallback failed ({fallback_failure})"
        ) from exc
    if historical_backfill:
        raise RBNZSourceUnavailableError(
            f"RBNZ B2 workbook failed ({primary_failure}); canonical HTML "
            f"fallback at {document.source_uri} is a recent summary and cannot "
            "satisfy a historical backfill"
        )
    return (document,)


def build_official_adapters(
    *,
    registry: MarketRegistry | None = None,
    repository: MarketRepository | None = None,
    backfill: bool = False,
    clock=None,
) -> tuple[FunctionalCanonicalAdapter, ...]:
    """Construct every deployable official adapter.

    Licensed and tenant connectors are declared by packs but are not invented
    here. SONIA is the one restricted official publication we ingest for
    derived computation; its pack policy prevents raw-value redistribution.
    """

    markets = registry or default_registry()
    now = (clock() if clock is not None else datetime.now(UTC)).astimezone(UTC)
    recent_start = now.date() - timedelta(days=45)
    configured_start = date.fromisoformat(
        os.getenv("SEICHE_CANONICAL_START", "2018-01-01")
    )
    start = configured_start if backfill else recent_start
    adapters: list[FunctionalCanonicalAdapter] = []

    def add(
        market_id: str,
        adapter_id: str,
        source: str,
        fetcher,
        parser,
        timeout=60.0,
        availability_check=None,
    ):
        adapters.append(
            FunctionalCanonicalAdapter(
                pack=markets.get(market_id),
                adapter_id=adapter_id,
                source=source,
                fetcher=fetcher,
                parser=parser,
                repository=repository,
                clock=clock,
                timeout_seconds=timeout,
                historical_backfill=backfill,
                availability_check=availability_check,
            )
        )

    fred_base = "https://fred.stlouisfed.org/graph/fredgraph.csv"
    fred_daily = {
        "US.FED.IORB": "IORB",
        "US.NYFED.SOFR": "SOFR",
        "US.NYFED.EFFR": "EFFR",
        "US.FED.POLICY_CEILING": "DFEDTARU",
        "US.TREASURY.TBILL_3M": "DTB3",
        "US.CP.NONFINANCIAL_3M": "DCPN3M",
    }

    async def fetch_fred_daily(client):
        return await get_documents(
            client,
            (
                (
                    instrument,
                    fred_base,
                    {"id": series, "cosd": start.isoformat()},
                )
                for instrument, series in fred_daily.items()
            ),
        )

    async def fetch_fred_weekly(client):
        return await get_documents(
            client,
            (
                (
                    "US.FED.RESERVE_BALANCES",
                    fred_base,
                    {"id": "WRESBAL", "cosd": start.isoformat()},
                ),
            ),
        )

    add("US-USD", "fred_daily", "fred", fetch_fred_daily, parse_fred_csv)
    add("US-USD", "fred_weekly", "fred", fetch_fred_weekly, parse_fred_csv)

    async def fetch_nyfed_rates(client):
        end = now.date().isoformat()
        return await get_documents(
            client,
            (
                (
                    "nyfed_secured_rates",
                    "https://markets.newyorkfed.org/api/rates/secured/all/search.json",
                    {"startDate": start.isoformat(), "endDate": end},
                ),
            ),
        )

    async def fetch_nyfed_unsecured_rates(client):
        end = now.date().isoformat()
        return await get_documents(
            client,
            (
                (
                    "nyfed_unsecured_rates",
                    "https://markets.newyorkfed.org/api/rates/unsecured/all/search.json",
                    {"startDate": start.isoformat(), "endDate": end},
                ),
            ),
        )

    async def fetch_nyfed_facilities(client):
        # The public endpoint accepts at most 500 results (1,000/1,200 return
        # HTTP 400). This reaches roughly one year of standing-facility
        # operations at current cadence and fails neither the US pack nor its
        # siblings during the one-time canonical import.
        count = 500 if backfill else 120
        return await get_documents(
            client,
            (
                (
                    "nyfed_repo_operations",
                    f"https://markets.newyorkfed.org/api/rp/repo/all/results/last/{count}.json",
                    None,
                ),
            ),
        )

    add("US-USD", "nyfed_rates", "nyfed_rates", fetch_nyfed_rates, parse_nyfed_rates)
    add(
        "US-USD",
        "nyfed_unsecured_rates",
        "nyfed_unsecured_rates",
        fetch_nyfed_unsecured_rates,
        parse_nyfed_unsecured_rates,
    )
    add(
        "US-USD",
        "nyfed_facilities",
        "nyfed_facilities",
        fetch_nyfed_facilities,
        parse_nyfed_facilities,
    )

    async def fetch_fiscaldata(client):
        uri = (
            "https://api.fiscaldata.treasury.gov/services/api/fiscal_service/"
            "v1/accounting/dts/operating_cash_balance"
        )
        common = {
            "filter": (
                f"record_date:gte:{start.isoformat()},account_type:in:("
                "Treasury General Account (TGA) Opening Balance,"
                "Treasury General Account (TGA),Federal Reserve Account)"
            ),
            "fields": "record_date,account_type,open_today_bal",
            "sort": "record_date",
            "page[size]": 1000,
        }
        documents = []
        page = 1
        total_pages = 1
        while page <= total_pages:
            response = await client.get(uri, params={**common, "page[number]": page})
            response.raise_for_status()
            payload = response.json()
            total_pages = min(
                int((payload.get("meta") or {}).get("total-pages", 1)), 20
            )
            documents.append(
                FetchedDocument(
                    str(response.url),
                    "application/json",
                    response.content,
                    f"tga-page-{page}",
                )
            )
            page += 1
        return tuple(documents)

    add("US-USD", "fiscaldata", "fiscaldata", fetch_fiscaldata, parse_fiscal_tga)

    ecb_base = "https://data-api.ecb.europa.eu/service/data"

    def ecb_fetcher(series: dict[str, str]):
        async def fetch(client):
            return await get_documents(
                client,
                (
                    (
                        instrument,
                        f"{ecb_base}/{remote}",
                        {"format": "csvdata", "startPeriod": start.isoformat()},
                    )
                    for instrument, remote in series.items()
                ),
            )

        return fetch

    add(
        "EA-EUR",
        "ecb_benchmark",
        "ecb_benchmark",
        ecb_fetcher(
            {
                "EA.ECB.ESTR": "EST/B.EU000A2X2A25.WT",
                "EA.ECB.ESTR_VOLUME": "EST/B.EU000A2X2A25.TT",
            }
        ),
        parse_ecb_csv,
    )
    add(
        "EA-EUR",
        "ecb_policy",
        "ecb_policy",
        ecb_fetcher(
            {
                "EA.ECB.DFR": "FM/D.U2.EUR.4F.KR.DFR.LEV",
                "EA.ECB.MRO": "FM/D.U2.EUR.4F.KR.MRR_FR.LEV",
                "EA.ECB.MLF": "FM/D.U2.EUR.4F.KR.MLFR.LEV",
            }
        ),
        parse_ecb_csv,
    )
    add(
        "EA-EUR",
        "ecb_liquidity",
        "ecb_liquidity",
        ecb_fetcher({"EA.ECB.EXCESS_LIQUIDITY": "ILM/D.U2.C.EXLIQ.U2.EUR"}),
        parse_ecb_csv,
    )

    boe_base = (
        "https://www.bankofengland.co.uk/boeapps/database/_iadb-fromshowcolumns.asp"
    )

    def boe_fetcher(instrument: str, code: str):
        async def fetch(client):
            return await get_documents(
                client,
                (
                    (
                        instrument,
                        boe_base,
                        {
                            "csv.x": "yes",
                            "Datefrom": start.strftime("%d/%b/%Y"),
                            "Dateto": "now",
                            "SeriesCodes": code,
                            "CSVF": "TN",
                            "UsingCodes": "Y",
                            "VPD": "Y",
                            "VFD": "N",
                        },
                    ),
                ),
            )

        return fetch

    add(
        "UK-GBP",
        "boe_sonia",
        "boe_sonia",
        boe_fetcher("GB.BOE.SONIA", "IUDSOIA"),
        parse_boe_csv,
    )
    add(
        "UK-GBP",
        "boe_policy",
        "boe_policy",
        boe_fetcher("GB.BOE.BANK_RATE", "IUDBEDR"),
        parse_boe_csv,
    )

    rba_f1 = "https://www.rba.gov.au/statistics/tables/xls/f01d.xlsx"
    rba_a2 = "https://www.rba.gov.au/statistics/tables/xls/a02hist.xlsx"

    async def fetch_rba_cash(client):
        return await get_documents(client, (("rba_cash", rba_f1, None),))

    async def fetch_rba_policy(client):
        return await get_documents(
            client,
            (("rba_policy_daily", rba_f1, None), ("rba_policy_changes", rba_a2, None)),
        )

    add("AU-AUD", "rba_cash", "rba_cash", fetch_rba_cash, parse_rba_workbook, 90)
    add("AU-AUD", "rba_policy", "rba_policy", fetch_rba_policy, parse_rba_workbook, 90)

    boj_base = "https://www.stat-search.boj.or.jp/ssi/mtshtml/csv"

    async def fetch_boj_rates(client):
        return await get_documents(
            client,
            (
                ("JP.BOJ.TONA", f"{boj_base}/fm01_d_1_en.csv", None),
                ("JP.BOJ.BASIC_LOAN", f"{boj_base}/ir01_d_1_en.csv", None),
            ),
        )

    async def fetch_boj_accounts(client):
        return await get_documents(
            client,
            (("JP.BOJ.CURRENT_ACCOUNTS", f"{boj_base}/md01_m_1_en.csv", None),),
        )

    add("JP-JPY", "boj_rates", "boj_rates", fetch_boj_rates, parse_boj_csv)
    add("JP-JPY", "boj_accounts", "boj_accounts", fetch_boj_accounts, parse_boj_csv)

    bok_end = now.astimezone(ZoneInfo("Asia/Seoul")).date()
    bok_start = configured_start if backfill else bok_end - timedelta(days=45)

    def bok_ecos_fetcher(series: dict[str, tuple[str, str]]):
        async def fetch(client):
            api_key = _bok_ecos_api_key()
            documents: list[FetchedDocument] = []
            for instrument, (stat_code, item_code) in series.items():
                first_row = 1
                total_rows: int | None = None
                while total_rows is None or first_row <= total_rows:
                    last_row = first_row + _BOK_ECOS_PAGE_SIZE - 1
                    uri = (
                        f"{_BOK_ECOS_BASE_URI}/{api_key}/json/en/"
                        f"{first_row}/{last_row}/{stat_code}/D/"
                        f"{bok_start:%Y%m%d}/{bok_end:%Y%m%d}/{item_code}"
                    )
                    request_fault: str | None = None
                    try:
                        response = await client.get(uri, follow_redirects=False)
                    except Exception as exc:  # noqa: BLE001 - credential boundary
                        request_fault = type(exc).__name__
                    if request_fault is not None:
                        # Raise after leaving the except block so the original
                        # credential-bearing request exception is not retained
                        # as __context__ on the public-safe fault.
                        raise BOKECOSSourceError(
                            f"BOK ECOS request failed ({request_fault})"
                        ) from None
                    if not 200 <= response.status_code < 300:
                        # Do not call raise_for_status(): httpx includes the full
                        # request URL, and ECOS puts its credential in that path.
                        raise BOKECOSSourceError(
                            f"BOK ECOS request failed with HTTP {response.status_code}"
                        )
                    if api_key.encode() in response.content:
                        raise BOKECOSSourceError(
                            "BOK ECOS response rejected because it echoed the API credential"
                        )
                    response_fault: str | None = None
                    try:
                        search, rows = _bok_ecos_search_payload(
                            response.content,
                            reject_credential=api_key,
                        )
                    except Exception as exc:  # noqa: BLE001 - credential boundary
                        response_fault = _redact_bok_ecos_credential(
                            f"{type(exc).__name__}: {exc}", api_key
                        )
                    if response_fault is not None:
                        raise BOKECOSSourceError(
                            f"BOK ECOS response validation failed: {response_fault}"
                        ) from None
                    try:
                        reported_total = int(search["list_total_count"])
                    except (KeyError, TypeError, ValueError) as exc:
                        raise ValueError(
                            "BOK ECOS response has no valid list_total_count"
                        ) from exc
                    if total_rows is None:
                        total_rows = reported_total
                    elif reported_total != total_rows:
                        raise ValueError(
                            "BOK ECOS list_total_count changed during pagination"
                        )
                    if total_rows < 1 or not rows:
                        raise ValueError(
                            f"BOK ECOS returned no rows for {instrument!r}"
                        )
                    documents.append(
                        FetchedDocument(
                            _redact_bok_ecos_credential(response.url, api_key),
                            "application/json",
                            response.content,
                            instrument,
                        )
                    )
                    first_row += len(rows)
            return tuple(documents)

        return fetch

    add(
        "KR-KRW",
        "bok_ecos_policy",
        "bok_ecos",
        bok_ecos_fetcher({"KR.BOK.BASE_RATE": _BOK_ECOS_SERIES["KR.BOK.BASE_RATE"]}),
        parse_bok_ecos,
        availability_check=_require_bok_ecos_api_key,
    )
    add(
        "KR-KRW",
        "bok_ecos_money_market",
        "bok_ecos",
        bok_ecos_fetcher(
            {"KR.BOK.CALL_OVERNIGHT_ALL": _BOK_ECOS_SERIES["KR.BOK.CALL_OVERNIGHT_ALL"]}
        ),
        parse_bok_ecos,
        availability_check=_require_bok_ecos_api_key,
    )

    async def fetch_cfets(client):
        end = now.date()
        window_start = start if backfill else max(start, end - timedelta(days=31))
        fdr = await client.get(
            "https://www.chinamoney.com.cn/r/cms/www/chinamoney/data/currency/fdr-chrt.csv",
            headers={"Referer": "https://www.chinamoney.com.cn/english/bmkfrr/"},
        )
        fdr.raise_for_status()
        documents = [
            FetchedDocument(
                str(fdr.url),
                "text/csv",
                fdr.content,
                "CN.CFETS.DR007",
            )
        ]
        # CFETS states and enforces a one-year maximum per historical query.
        # Use 360 inclusive days to avoid leap-year boundary ambiguity.
        shibor_start = max(window_start, date(2007, 1, 4))
        for chunk_start, chunk_end in bounded_date_windows(
            shibor_start,
            end,
            maximum_days=360,
        ):
            shibor = await client.get(
                "https://www.chinamoney.com.cn/ags/ms/cm-u-bk-shibor/ShiborHis",
                params={
                    "lang": "en",
                    "startDate": chunk_start.isoformat(),
                    "endDate": chunk_end.isoformat(),
                },
                headers={"Referer": "https://www.chinamoney.com.cn/english/bmkshibor/"},
            )
            shibor.raise_for_status()
            records = shibor.json().get("records") or []
            if records:
                documents.append(
                    FetchedDocument(
                        str(shibor.url),
                        "application/json",
                        shibor.content,
                        "CN.CFETS.SHIBOR_ON",
                    )
                )
        return tuple(documents)

    add("CN-CNY", "cfets_rates", "cfets_rates", fetch_cfets, parse_cfets_rates, 90)

    add(
        "HK-HKD",
        "hkma_official",
        "hkma_official",
        _fetch_hkma_documents,
        parse_hkma_json,
    )

    async def fetch_rbi(client):
        return await get_documents(
            client,
            (
                (
                    "rbi_mmo",
                    "https://www.rbi.org.in/Scripts/BS_ViewMMO.aspx/Statistics.aspx",
                    None,
                ),
                ("rbi_home", "https://www.rbi.org.in/", None),
            ),
        )

    add("IN-INR", "rbi_official", "rbi_official", fetch_rbi, parse_rbi_html, 90)

    def rbnz_fetcher(label: str):
        async def fetch(client):
            return await _fetch_rbnz_document(
                client,
                label,
                historical_backfill=backfill,
            )

        return fetch

    add(
        "NZ-NZD",
        "rbnz_policy",
        "rbnz_policy",
        rbnz_fetcher("rbnz_policy"),
        parse_rbnz,
        90,
        availability_check=_require_rbnz_access_approval,
    )
    add(
        "NZ-NZD",
        "rbnz_wholesale",
        "rbnz_wholesale",
        rbnz_fetcher("rbnz_wholesale"),
        parse_rbnz,
        90,
        availability_check=_require_rbnz_access_approval,
    )

    mas_start_year = configured_start.year if backfill else now.year

    async def fetch_mas_sora(client):
        return await _mas_documents(
            client,
            label="mas_sora",
            columns=(13, 18),
            start_year=mas_start_year,
            end_year=now.year,
            end_month=now.month,
        )

    async def fetch_mas_rates(client):
        return await _mas_documents(
            client,
            label="mas_rates",
            columns=(10, 11),
            start_year=mas_start_year,
            end_year=now.year,
            end_month=now.month,
        )

    add("SG-SGD", "mas_sora", "mas_sora", fetch_mas_sora, parse_mas_sora, 120)
    add("SG-SGD", "mas_rates", "mas_rates", fetch_mas_rates, parse_mas_rates, 120)

    return tuple(adapters)


PRODUCTION_ADAPTER_KEYS = frozenset(
    {
        ("US-USD", "fred_daily"),
        ("US-USD", "fred_weekly"),
        ("US-USD", "nyfed_rates"),
        ("US-USD", "nyfed_unsecured_rates"),
        ("US-USD", "nyfed_facilities"),
        ("US-USD", "fiscaldata"),
        ("EA-EUR", "ecb_benchmark"),
        ("EA-EUR", "ecb_policy"),
        ("EA-EUR", "ecb_liquidity"),
        ("UK-GBP", "boe_sonia"),
        ("UK-GBP", "boe_policy"),
        ("JP-JPY", "boj_rates"),
        ("JP-JPY", "boj_accounts"),
        ("KR-KRW", "bok_ecos_policy"),
        ("KR-KRW", "bok_ecos_money_market"),
        ("CN-CNY", "cfets_rates"),
        ("HK-HKD", "hkma_official"),
        ("IN-INR", "rbi_official"),
        ("AU-AUD", "rba_cash"),
        ("AU-AUD", "rba_policy"),
        ("NZ-NZD", "rbnz_policy"),
        ("NZ-NZD", "rbnz_wholesale"),
        ("SG-SGD", "mas_sora"),
        ("SG-SGD", "mas_rates"),
    }
)
