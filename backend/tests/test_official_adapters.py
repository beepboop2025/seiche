from __future__ import annotations

import asyncio
import gzip
import hashlib
import io
import json
import os
import re
import urllib.parse
from datetime import UTC, date, datetime
from decimal import Decimal

import httpx
import pytest

from seiche import store
from seiche.collectors import CollectorRunStatus, CollectorSupervisor
from seiche.domain.observation import (
    RATE_ROLES,
    CanonicalUnit,
    ConnectorClassification,
    DayCountConvention,
    QualityState,
    RateCompounding,
    RedistributionStatus,
    SemanticRole,
    evidence_sha256,
)
from seiche.markets.registry import default_registry
from seiche.repository import SQLiteMarketRepository
from seiche.sources.base import (
    ObservationBatch,
    RawCapture,
    SourcePolicyUnavailableError,
)
from seiche.sources.canonical import (
    FetchedDocument,
    FunctionalCanonicalAdapter,
    ParsedPoint,
)
from seiche.sources import official
from seiche.sources.official import (
    PRODUCTION_ADAPTER_KEYS,
    RBNZSourceUnavailableError,
    bounded_date_windows,
    build_official_adapters,
    parse_nyfed_rates,
    parse_nyfed_unsecured_rates,
    parse_rbnz,
)


def test_bounded_date_windows_are_complete_non_overlapping_and_inclusive() -> None:
    windows = bounded_date_windows(
        date(2024, 1, 1),
        date(2026, 1, 5),
        maximum_days=360,
    )

    assert windows[0][0] == date(2024, 1, 1)
    assert windows[-1][1] == date(2026, 1, 5)
    assert all((end - start).days < 360 for start, end in windows)
    assert all(
        current[1].toordinal() + 1 == following[0].toordinal()
        for current, following in zip(windows, windows[1:])
    )


def test_nyfed_sofr_median_uses_percent_rate_and_binds_field_date_lineage() -> None:
    document = FetchedDocument(
        "https://markets.newyorkfed.org/api/rates/secured/all/search.json",
        "application/json",
        b'{"refRates":[{"type":"SOFR","effectiveDate":"2026-08-07",'
        b'"percentRate":5.31,"percentPercentile25":4.87,'
        b'"percentPercentile99":5.45,"volumeInBillions":2100}]}',
        "nyfed_secured_rates",
    )

    points = parse_nyfed_rates(document)
    by_instrument = {point.instrument_id: point for point in points}
    median = by_instrument["US.NYFED.SOFR_MEDIAN"]

    assert median.raw_value == Decimal("5.31")
    assert median.raw_value != Decimal("4.87")
    assert re.fullmatch(
        r"nyfed:percentRate:2026-08-07:unrevised-[0-9a-f]{16}",
        str(median.revision_id),
    )
    assert re.fullmatch(
        r"nyfed:percentPercentile99:2026-08-07:unrevised-[0-9a-f]{16}",
        str(by_instrument["US.NYFED.SOFR_P99"].revision_id),
    )
    assert re.fullmatch(
        r"nyfed:volumeInBillions:2026-08-07:unrevised-[0-9a-f]{16}",
        str(by_instrument["US.NYFED.SOFR_VOLUME"].revision_id),
    )


def test_nyfed_revision_indicator_cannot_replace_field_and_event_lineage() -> None:
    document = FetchedDocument(
        "https://markets.newyorkfed.org/api/rates/secured/all/search.json",
        "application/json",
        b'{"refRates":[{"type":"SOFR","effectiveDate":"2026-08-06",'
        b'"percentRate":5.3,"percentPercentile25":5.2,'
        b'"percentPercentile99":5.4,"volumeInBillions":2000,'
        b'"revisionIndicator":"R1"}]}',
        "nyfed_secured_rates",
    )

    points = parse_nyfed_rates(document)

    assert len({point.revision_id for point in points}) == 3
    assert all(
        re.fullmatch(
            r"nyfed:(?:percentRate|percentPercentile99|volumeInBillions):"
            r"2026-08-06:R1-[0-9a-f]{16}",
            str(point.revision_id),
        )
        for point in points
    )


def test_nyfed_changed_unflagged_row_gets_distinct_revision_lineage() -> None:
    first = FetchedDocument(
        "https://markets.newyorkfed.org/api/rates/secured/all/search.json",
        "application/json",
        b'{"refRates":[{"type":"SOFR","effectiveDate":"2026-08-06",'
        b'"percentRate":5.30,"percentPercentile99":5.40,'
        b'"volumeInBillions":2000}]}',
        "nyfed_secured_rates",
    )
    changed = FetchedDocument(
        first.source_uri,
        first.media_type,
        b'{"refRates":[{"type":"SOFR","effectiveDate":"2026-08-06",'
        b'"percentRate":5.31,"percentPercentile99":5.40,'
        b'"volumeInBillions":2000}]}',
        first.label,
    )

    first_median = next(
        point
        for point in parse_nyfed_rates(first)
        if point.instrument_id == "US.NYFED.SOFR_MEDIAN"
    )
    changed_median = next(
        point
        for point in parse_nyfed_rates(changed)
        if point.instrument_id == "US.NYFED.SOFR_MEDIAN"
    )

    assert first_median.revision_id != changed_median.revision_id
    assert str(first_median.revision_id).startswith(
        "nyfed:percentRate:2026-08-06:unrevised-"
    )


def test_nyfed_sofrai_parser_maps_all_horizons_and_index_with_same_day_clock() -> None:
    document = FetchedDocument(
        "https://markets.newyorkfed.org/api/rates/secured/all/search.json",
        "application/json",
        b'{"refRates":[{"effectiveDate":"2026-08-21","type":"SOFRAI",'
        b'"average30day":3.64319,"average90day":3.63920,'
        b'"average180day":3.66017,"index":1.25565902,'
        b'"revisionIndicator":"R1"}]}',
        "nyfed_secured_rates",
    )

    points = parse_nyfed_rates(document)
    by_instrument = {point.instrument_id: point for point in points}

    assert {
        instrument: point.raw_value for instrument, point in by_instrument.items()
    } == {
        "US.NYFED.SOFR_AVERAGE_30D": Decimal("3.64319"),
        "US.NYFED.SOFR_AVERAGE_90D": Decimal("3.63920"),
        "US.NYFED.SOFR_AVERAGE_180D": Decimal("3.66017"),
        "US.NYFED.SOFR_INDEX": Decimal("1.25565902"),
    }
    assert all(
        point.source_publication_time == datetime(2026, 8, 21, 12, tzinfo=UTC)
        for point in points
    )
    assert all(
        re.fullmatch(
            r"nyfed:SOFRAI:(?:average30day|average90day|average180day|index):"
            r"2026-08-21:R1-[0-9a-f]{16}",
            str(point.revision_id),
        )
        for point in points
    )
    evidence = json.loads(by_instrument["US.NYFED.SOFR_INDEX"].row_evidence)
    assert evidence["label"] == "index"
    assert evidence["row"]["average180day"] == 3.66017


def test_nyfed_sofrai_parser_never_fills_missing_horizon() -> None:
    document = FetchedDocument(
        "https://markets.newyorkfed.org/api/rates/secured/all/search.json",
        "application/json",
        b'{"refRates":[{"effectiveDate":"2026-08-21","type":"SOFRAI",'
        b'"average30day":null,"average90day":3.63920,'
        b'"average180day":"","index":1.25565902}]}',
        "nyfed_secured_rates",
    )

    points = parse_nyfed_rates(document)

    assert [point.instrument_id for point in points] == [
        "US.NYFED.SOFR_AVERAGE_90D",
        "US.NYFED.SOFR_INDEX",
    ]


def test_nyfed_sofrai_parser_rejects_duplicate_field_date_identity() -> None:
    row = (
        b'{"effectiveDate":"2026-08-21","type":"SOFRAI",'
        b'"average30day":3.64319,"average90day":3.63920,'
        b'"average180day":3.66017,"index":1.25565902}'
    )
    document = FetchedDocument(
        "https://markets.newyorkfed.org/api/rates/secured/all/search.json",
        "application/json",
        b'{"refRates":[' + row + b"," + row + b"]}",
        "nyfed_secured_rates",
    )

    with pytest.raises(ValueError, match="duplicate SOFRAI average30day"):
        parse_nyfed_rates(document)


def test_nyfed_sofrai_changed_unflagged_row_gets_new_content_lineage() -> None:
    first = FetchedDocument(
        "https://markets.newyorkfed.org/api/rates/secured/all/search.json",
        "application/json",
        b'{"refRates":[{"effectiveDate":"2026-08-21","type":"SOFRAI",'
        b'"average30day":3.64319,"average90day":3.63920,'
        b'"average180day":3.66017,"index":1.25565902}]}',
        "nyfed_secured_rates",
    )
    changed = FetchedDocument(
        first.source_uri,
        first.media_type,
        first.payload.replace(b'"average30day":3.64319', b'"average30day":3.64320'),
        first.label,
    )

    first_average = next(
        point
        for point in parse_nyfed_rates(first)
        if point.instrument_id == "US.NYFED.SOFR_AVERAGE_30D"
    )
    changed_average = next(
        point
        for point in parse_nyfed_rates(changed)
        if point.instrument_id == "US.NYFED.SOFR_AVERAGE_30D"
    )

    assert first_average.revision_id != changed_average.revision_id
    assert str(first_average.revision_id).startswith(
        "nyfed:SOFRAI:average30day:2026-08-21:unrevised-"
    )


def test_nyfed_unsecured_parser_maps_canonical_effr_and_obfr_distribution() -> None:
    document = FetchedDocument(
        "https://markets.newyorkfed.org/api/rates/unsecured/all/search.json",
        "application/json",
        b'{"refRates":['
        b'{"type":"OBFR","effectiveDate":"2026-08-20",'
        b'"percentRate":3.63,"percentPercentile1":3.53,'
        b'"percentPercentile25":3.62,"percentPercentile75":3.63,'
        b'"percentPercentile99":3.68,"volumeInBillions":229,'
        b'"revisionIndicator":""},'
        b'{"type":"EFFR","effectiveDate":"2026-08-20",'
        b'"percentRate":3.63,"percentPercentile1":3.60,'
        b'"percentPercentile25":3.62,"percentPercentile75":3.63,'
        b'"percentPercentile99":3.69,"volumeInBillions":102,'
        b'"revisionIndicator":"R1"}]}',
        "nyfed_unsecured_rates",
    )

    points = parse_nyfed_unsecured_rates(document)
    by_instrument = {point.instrument_id: point for point in points}

    assert [point.instrument_id for point in points] == [
        "US.NYFED.EFFR_MEDIAN",
        "US.NYFED.EFFR_P99",
        "US.NYFED.OBFR_MEDIAN",
        "US.NYFED.OBFR_P99",
    ]
    assert {
        instrument: point.raw_value for instrument, point in by_instrument.items()
    } == {
        "US.NYFED.EFFR_MEDIAN": Decimal("3.63"),
        "US.NYFED.EFFR_P99": Decimal("3.69"),
        "US.NYFED.OBFR_MEDIAN": Decimal("3.63"),
        "US.NYFED.OBFR_P99": Decimal("3.68"),
    }
    assert re.fullmatch(
        r"nyfed:EFFR:percentRate:2026-08-20:R1-[0-9a-f]{16}",
        str(by_instrument["US.NYFED.EFFR_MEDIAN"].revision_id),
    )
    assert re.fullmatch(
        r"nyfed:OBFR:percentPercentile99:2026-08-20:unrevised-[0-9a-f]{16}",
        str(by_instrument["US.NYFED.OBFR_P99"].revision_id),
    )
    evidence = json.loads(by_instrument["US.NYFED.EFFR_MEDIAN"].row_evidence)
    assert evidence["label"] == "EFFR.percentRate"
    assert evidence["row"]["volumeInBillions"] == 102


def test_nyfed_unsecured_parser_is_order_stable_and_never_fills_missing_tail() -> None:
    effr = (
        b'{"type":"EFFR","effectiveDate":"2026-08-20",'
        b'"percentRate":3.63,"percentPercentile99":null}'
    )
    obfr = (
        b'{"type":"OBFR","effectiveDate":"2026-08-20",'
        b'"percentRate":3.64,"percentPercentile99":3.68}'
    )

    def parse(rows: bytes):
        return parse_nyfed_unsecured_rates(
            FetchedDocument(
                "https://markets.newyorkfed.org/api/rates/unsecured/all/search.json",
                "application/json",
                b'{"refRates":[' + rows + b"]}",
                "nyfed_unsecured_rates",
            )
        )

    forward = parse(effr + b"," + obfr)
    reversed_rows = parse(obfr + b"," + effr)

    assert forward == reversed_rows
    assert [point.instrument_id for point in forward] == [
        "US.NYFED.EFFR_MEDIAN",
        "US.NYFED.OBFR_MEDIAN",
        "US.NYFED.OBFR_P99",
    ]
    assert all(point.instrument_id != "US.NYFED.EFFR_P99" for point in forward)


def test_nyfed_unsecured_parser_rejects_duplicate_distribution_identity() -> None:
    row = (
        b'{"type":"EFFR","effectiveDate":"2026-08-20",'
        b'"percentRate":3.63,"percentPercentile99":3.69}'
    )
    document = FetchedDocument(
        "https://markets.newyorkfed.org/api/rates/unsecured/all/search.json",
        "application/json",
        b'{"refRates":[' + row + b"," + row + b"]}",
        "nyfed_unsecured_rates",
    )

    with pytest.raises(ValueError, match="duplicate EFFR percentRate"):
        parse_nyfed_unsecured_rates(document)


def test_every_production_adapter_is_pack_declared_without_network_io(
    tmp_path, monkeypatch
) -> None:
    monkeypatch.setattr(store, "DB_PATH", tmp_path / "adapters.sqlite")
    adapters = build_official_adapters(
        repository=SQLiteMarketRepository(),
        clock=lambda: datetime(2026, 8, 9, tzinfo=UTC),
    )
    keys = {(item.market_id, item.adapter_id) for item in adapters}

    assert keys == PRODUCTION_ADAPTER_KEYS
    registry = default_registry()
    assert all(
        adapter.adapter_id in registry.get(adapter.market_id).adapter_map
        for adapter in adapters
    )


def _official_adapter(
    adapter_id: str, *, backfill: bool = False
) -> FunctionalCanonicalAdapter:
    adapters = build_official_adapters(
        repository=SQLiteMarketRepository(),
        backfill=backfill,
        clock=lambda: datetime(2026, 8, 11, tzinfo=UTC),
    )
    return next(item for item in adapters if item.adapter_id == adapter_id)


def test_nyfed_unsecured_adapter_registration_and_pack_contract() -> None:
    pack = default_registry().get("US-USD")
    spec = pack.adapter_map["nyfed_unsecured_rates"]
    instruments = {
        item.instrument_id: item
        for item in pack.instruments
        if item.source_adapter_id == "nyfed_unsecured_rates"
    }

    assert spec.classification is ConnectorClassification.OFFICIAL_OPEN
    assert spec.redistribution_status is RedistributionStatus.ALLOWED
    assert spec.expected_cadence == "P1D"
    assert spec.publication_clock.business_day_lag == 1
    assert spec.publication_clock.local_time is not None
    assert spec.publication_clock.local_time.isoformat() == "09:00:00"
    assert set(instruments) == {
        "US.NYFED.EFFR_MEDIAN",
        "US.NYFED.EFFR_P99",
        "US.NYFED.OBFR_MEDIAN",
        "US.NYFED.OBFR_P99",
    }
    assert {
        instrument: item.semantic_role for instrument, item in instruments.items()
    } == {
        "US.NYFED.EFFR_MEDIAN": SemanticRole.RATE_MEDIAN,
        "US.NYFED.EFFR_P99": SemanticRole.RATE_P99,
        "US.NYFED.OBFR_MEDIAN": SemanticRole.RATE_MEDIAN,
        "US.NYFED.OBFR_P99": SemanticRole.RATE_P99,
    }
    assert all(item.source_unit == "percent" for item in instruments.values())
    assert all(
        item.canonical_unit is CanonicalUnit.BASIS_POINTS
        and item.value_multiplier == Decimal("100")
        for item in instruments.values()
    )
    assert (
        pack.instrument_map["US.NYFED.SOFR_MEDIAN"].source_adapter_id == "nyfed_rates"
    )
    adapter = _official_adapter("nyfed_unsecured_rates")
    assert adapter.source == "nyfed_unsecured_rates"
    assert adapter.parser is parse_nyfed_unsecured_rates


def test_nyfed_sofrai_pack_contract_has_honest_rate_and_index_semantics() -> None:
    pack = default_registry().get("US-USD")
    average_roles = {
        "US.NYFED.SOFR_AVERAGE_30D": SemanticRole.COMPOUNDED_OVERNIGHT_AVERAGE_30D,
        "US.NYFED.SOFR_AVERAGE_90D": SemanticRole.COMPOUNDED_OVERNIGHT_AVERAGE_90D,
        "US.NYFED.SOFR_AVERAGE_180D": SemanticRole.COMPOUNDED_OVERNIGHT_AVERAGE_180D,
    }

    for instrument_id, role in average_roles.items():
        instrument = pack.instrument_map[instrument_id]
        assert instrument.source_adapter_id == "nyfed_rates"
        assert instrument.semantic_role is role
        assert role in RATE_ROLES
        assert instrument.source_unit == "percent"
        assert instrument.canonical_unit is CanonicalUnit.BASIS_POINTS
        assert instrument.value_multiplier == Decimal("100")
        assert instrument.rate_compounding is RateCompounding.COMPOUNDED
        assert instrument.day_count is DayCountConvention.ACT_360

    index = pack.instrument_map["US.NYFED.SOFR_INDEX"]
    assert index.source_adapter_id == "nyfed_rates"
    assert index.semantic_role is SemanticRole.COMPOUNDED_OVERNIGHT_RATE_INDEX
    assert index.semantic_role not in RATE_ROLES
    assert index.source_unit == "index points"
    assert index.canonical_unit is CanonicalUnit.INDEX_POINTS
    assert index.value_multiplier == Decimal("1")
    assert index.rate_compounding is None
    assert index.day_count is None


@pytest.mark.asyncio
async def test_nyfed_unsecured_fetcher_uses_bounded_official_search() -> None:
    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        return httpx.Response(
            200,
            headers={"content-type": "application/json; charset=utf-8"},
            json={
                "refRates": [
                    {
                        "type": "EFFR",
                        "effectiveDate": "2026-08-10",
                        "percentRate": 3.63,
                        "percentPercentile99": 3.69,
                    }
                ]
            },
        )

    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        documents = tuple(
            await _official_adapter("nyfed_unsecured_rates").fetcher(client)
        )

    assert len(requests) == 1
    assert requests[0].url.path == "/api/rates/unsecured/all/search.json"
    assert dict(requests[0].url.params) == {
        "startDate": "2026-06-27",
        "endDate": "2026-08-11",
    }
    assert len(documents) == 1
    assert documents[0].label == "nyfed_unsecured_rates"
    assert documents[0].media_type == "application/json"
    assert documents[0].source_uri == str(requests[0].url)


@pytest.mark.asyncio
async def test_nyfed_unsecured_adapter_preserves_raw_capture_and_row_clocks(
    tmp_path, monkeypatch
) -> None:
    monkeypatch.setattr(store, "DB_PATH", tmp_path / "nyfed-unsecured.sqlite")
    capture = datetime(2026, 8, 11, 14, tzinfo=UTC)
    payload = (
        b'{"refRates":[{"type":"EFFR","effectiveDate":"2026-08-10",'
        b'"percentRate":3.63,"percentPercentile99":3.69,'
        b'"revisionIndicator":""}]}'
    )
    document = FetchedDocument(
        "https://markets.newyorkfed.org/api/rates/unsecured/all/search.json"
        "?startDate=2026-08-10&endDate=2026-08-11",
        "application/json",
        payload,
        "nyfed_unsecured_rates",
    )

    async def fetcher(_client):
        return (document,)

    adapter = next(
        item
        for item in build_official_adapters(
            repository=SQLiteMarketRepository(),
            clock=lambda: capture,
        )
        if item.adapter_id == "nyfed_unsecured_rates"
    )
    adapter.fetcher = fetcher

    batch = await adapter.collect()
    by_instrument = {row.instrument_id: row for row in batch.observations}

    assert batch.raw_capture is not None
    assert batch.raw_capture.payload == payload
    assert batch.raw_capture.source_uri == document.source_uri
    assert batch.raw_capture.media_type == "application/json"
    assert by_instrument["US.NYFED.EFFR_MEDIAN"].value == Decimal("363")
    assert by_instrument["US.NYFED.EFFR_P99"].value == Decimal("369")
    assert all(row.knowledge_time == capture for row in batch.observations)
    assert all(
        row.source_publication_time == datetime(2026, 8, 11, 13, tzinfo=UTC)
        for row in batch.observations
    )
    assert all(row.quality is QualityState.VERIFIED for row in batch.observations)


@pytest.mark.asyncio
async def test_nyfed_sofrai_adapter_preserves_raw_capture_units_and_bitemporal_clocks(
    tmp_path, monkeypatch
) -> None:
    monkeypatch.setattr(store, "DB_PATH", tmp_path / "nyfed-sofrai.sqlite")
    capture = datetime(2026, 8, 21, 12, 30, tzinfo=UTC)
    payload = (
        b'{"refRates":[{"effectiveDate":"2026-08-21","type":"SOFRAI",'
        b'"average30day":3.64319,"average90day":3.63920,'
        b'"average180day":3.66017,"index":1.25565902,'
        b'"revisionIndicator":""}]}'
    )
    document = FetchedDocument(
        "https://markets.newyorkfed.org/api/rates/secured/all/search.json"
        "?startDate=2026-08-21&endDate=2026-08-21",
        "application/json",
        payload,
        "nyfed_secured_rates",
    )

    async def fetcher(_client):
        return (document,)

    adapter = next(
        item
        for item in build_official_adapters(
            repository=SQLiteMarketRepository(),
            clock=lambda: capture,
        )
        if item.adapter_id == "nyfed_rates"
    )
    adapter.fetcher = fetcher

    batch = await adapter.collect()
    by_instrument = {row.instrument_id: row for row in batch.observations}
    average = by_instrument["US.NYFED.SOFR_AVERAGE_30D"]
    index = by_instrument["US.NYFED.SOFR_INDEX"]

    assert batch.raw_capture is not None
    assert batch.raw_capture.payload == payload
    assert batch.raw_capture.source_uri == document.source_uri
    assert average.value == Decimal("364.31900")
    assert average.canonical_unit is CanonicalUnit.BASIS_POINTS
    assert average.rate_compounding is RateCompounding.COMPOUNDED
    assert average.day_count is DayCountConvention.ACT_360
    assert index.value == Decimal("1.25565902")
    assert index.canonical_unit is CanonicalUnit.INDEX_POINTS
    assert index.rate_compounding is None
    assert index.day_count is None
    assert all(
        row.event_time == datetime(2026, 8, 21, tzinfo=UTC)
        for row in batch.observations
    )
    assert all(
        row.source_publication_time == datetime(2026, 8, 21, 12, tzinfo=UTC)
        for row in batch.observations
    )
    assert all(row.knowledge_time == capture for row in batch.observations)
    assert all(row.quality is QualityState.VERIFIED for row in batch.observations)


def _rbnz_workbook_payload(*, include_b2: bool = True) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    if include_b2:
        worksheet.append(
            [
                "Date",
                "Official Cash Rate (OCR)",
                "Overnight Deposit Rate",
                "Overnight Reverse Repurchase Facility Rate",
            ]
        )
        worksheet.append([datetime(2026, 8, 10), 2.50, 2.50, 3.00])
    else:
        worksheet.append(["Date", "Unrelated series"])
        worksheet.append([datetime(2026, 8, 10), 9.99])
    payload = io.BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def _rbnz_current_workbook_payload(
    *,
    reverse_series_id: str = "INM.DD2.N",
    duplicate_reverse: bool = False,
    series_marker: str = "Series Id",
    add_legacy_candidate: bool = False,
) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append([None, "Cash rate", "Cash rate", "Cash rate"])
    worksheet.append(
        [
            None,
            "Official Cash Rate (OCR)",
            "Overnight Deposit Rate",
            "Overnight Reverse Repurchase Facility Rate",
        ]
    )
    worksheet.append(["Notes", None, None, None])
    worksheet.append(["Unit", "%pa", "%pa", "%pa"])
    worksheet.append([series_marker, "INM.DP1.N", "INM.DD1.N", reverse_series_id])
    worksheet.append([datetime(2026, 8, 10), 2.50, 2.50, 3.00])
    if duplicate_reverse:
        worksheet.cell(
            row=2,
            column=5,
            value="Overnight Reverse Repurchase Facility Rate",
        )
        worksheet.cell(row=5, column=5, value="INM.DD2.N")
        worksheet.cell(row=6, column=5, value=3.00)
    workbook.create_sheet("Table Description").append(
        ["Table", "Daily wholesale interest rates (% pa) - B2"]
    )
    workbook.create_sheet("Series Definitions").append(
        ["Group", "Series", "Series Id", "Unit", "Note"]
    )
    if add_legacy_candidate:
        legacy = workbook.create_sheet("Legacy")
        legacy.append(
            [
                "Date",
                "Official Cash Rate (OCR)",
                "Overnight Deposit Rate",
                "Overnight Reverse Repurchase Facility Rate",
            ]
        )
        legacy.append([datetime(2026, 8, 10), 9.99, 9.99, 9.99])
    payload = io.BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def _approve_rbnz_access(monkeypatch) -> None:
    monkeypatch.setattr(official, "_rbnz_access_today", lambda: date(2026, 8, 14))
    monkeypatch.setenv(official._RBNZ_ACCESS_APPROVAL_SHA256_ENV, "a" * 64)
    monkeypatch.setenv(
        official._RBNZ_ACCESS_APPROVAL_VALID_UNTIL_ENV,
        "2027-08-14",
    )


def _cfets_approval_payload(
    *,
    valid_until: str = "2027-08-14",
    overrides: dict[str, str] | None = None,
    extra_fields: dict[str, str] | None = None,
) -> bytes:
    fields = {
        **official._CFETS_APPROVAL_FIXED_FIELDS,
        "licence_evidence_path": str(
            official._CFETS_LICENCE_EVIDENCE_CANONICAL_PATH
        ),
        "licence_evidence_sha256": "c" * 64,
        "valid_until": valid_until,
    }
    fields.update(overrides or {})
    fields.update(extra_fields or {})
    return "".join(f"{key}={value}\n" for key, value in fields.items()).encode()


def _approve_cfets_access(
    tmp_path,
    monkeypatch,
    *,
    valid_until: str = "2027-08-14",
    overrides: dict[str, str] | None = None,
    extra_fields: dict[str, str] | None = None,
):
    evidence_path = tmp_path / "cfets-licence-evidence.pdf"
    evidence_payload = b"%PDF-1.7\n% reviewed CFETS permission fixture\n%%EOF\n"
    evidence_path.write_bytes(evidence_payload)
    evidence_path.chmod(0o640)
    monkeypatch.setattr(
        official,
        "_CFETS_LICENCE_EVIDENCE_CANONICAL_PATH",
        evidence_path,
    )
    scoped_overrides = {
        "licence_evidence_path": str(evidence_path),
        "licence_evidence_sha256": hashlib.sha256(evidence_payload).hexdigest(),
        **(overrides or {}),
    }
    approval_path = tmp_path / "cfets-approval.conf"
    payload = _cfets_approval_payload(
        valid_until=valid_until,
        overrides=scoped_overrides,
        extra_fields=extra_fields,
    )
    approval_path.write_bytes(payload)
    approval_path.chmod(0o640)
    artifact_sha256 = hashlib.sha256(payload).hexdigest()
    monkeypatch.setattr(official, "_CFETS_APPROVAL_CANONICAL_PATH", approval_path)
    monkeypatch.setattr(
        official,
        "_cfets_approval_expected_owner",
        lambda: (os.getuid(), os.getgid()),
    )
    monkeypatch.setattr(official, "_cfets_access_today", lambda: date(2026, 8, 14))
    monkeypatch.setenv(official._CFETS_APPROVAL_PATH_ENV, str(approval_path))
    monkeypatch.setenv(official._CFETS_APPROVAL_SHA256_ENV, artifact_sha256)
    return approval_path, artifact_sha256


def _cfets_success_response(request: httpx.Request) -> httpx.Response:
    if request.url.path.endswith("/fdr-settings.json"):
        return httpx.Response(
            200,
            request=request,
            json={
                "columns": list(official._CFETS_FDR_CHART_COLUMNS),
                "graphs": [
                    {"gid": graph_id, "title": graph_id}
                    for graph_id in official._CFETS_FDR_GRAPH_IDS
                ],
            },
        )
    if request.url.path.endswith("/fdr-chrt.csv"):
        return httpx.Response(
            200,
            request=request,
            content=b"2026-08-11,0,0,0,0,0,1.51,1.52,1.53\n",
        )
    start = date.fromisoformat(request.url.params["startDate"])
    end = date.fromisoformat(request.url.params["endDate"])
    fixture_day = date(2026, 8, 11)
    event_day = fixture_day if start <= fixture_day <= end else end
    return httpx.Response(
        200,
        request=request,
        json={"records": [{"showDateCN": event_day.isoformat(), "ON": "1.31"}]},
    )


class _CFETSFinalResponseProbe:
    """Expose fetch completion to the supervisor persistence regression."""

    market_id = "CN-CNY"
    adapter_id = "cfets_rates"

    def __init__(
        self,
        delegate: FunctionalCanonicalAdapter,
        client: httpx.AsyncClient,
    ) -> None:
        self.delegate = delegate
        self.client = client
        self.fetch_completed = False

    def check_availability(self) -> None:
        self.delegate.check_availability()

    async def collect(self) -> ObservationBatch:
        documents = tuple(await self.delegate.fetcher(self.client))
        self.fetch_completed = True
        payload = documents[0].payload
        captured_at = datetime(2026, 8, 14, tzinfo=UTC)
        return ObservationBatch(
            market_id=self.market_id,
            adapter_id=self.adapter_id,
            captured_at=captured_at,
            observations=(),
            raw_capture=RawCapture(
                market_id=self.market_id,
                adapter_id=self.adapter_id,
                captured_at=captured_at,
                source_uri=documents[0].source_uri,
                media_type=documents[0].media_type,
                payload=payload,
                evidence_hash=evidence_sha256(payload),
            ),
        )


class _CFETSPersistenceRecorder:
    def __init__(self) -> None:
        self.writes: list[object] = []

    def write(self, value: object) -> list[str]:
        self.writes.append(value)
        return []


async def _run_cfets_final_response_probe(
    delegate: FunctionalCanonicalAdapter,
    handler,
) -> tuple[
    _CFETSFinalResponseProbe,
    list,
    _CFETSPersistenceRecorder,
    _CFETSPersistenceRecorder,
    list[tuple],
]:
    raw_sink = _CFETSPersistenceRecorder()
    normalized_sink = _CFETSPersistenceRecorder()
    observation_writes: list[tuple] = []
    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        probe = _CFETSFinalResponseProbe(delegate, client)
        supervisor = CollectorSupervisor(
            raw_sink=raw_sink,
            normalized_sink=normalized_sink,
            observation_writer=lambda rows: observation_writes.append(rows) or len(rows),
            persistence_retry_limit=0,
        )
        supervisor.register(probe)
        runs = await supervisor.run_due(
            now=datetime(2026, 8, 14, tzinfo=UTC),
            force=True,
        )
    return probe, runs, raw_sink, normalized_sink, observation_writes


def _rbnz_mock_transport(handler) -> tuple[httpx.MockTransport, list[httpx.Request]]:
    requests: list[httpx.Request] = []

    def wrapped(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        response = handler(request)
        if response.is_stream_consumed:
            return httpx.Response(
                response.status_code,
                headers=response.headers,
                stream=httpx.ByteStream(response.content),
            )
        return response

    return httpx.MockTransport(wrapped), requests


def test_connector_owned_retry_policies_are_not_multiplied_by_supervisor() -> None:
    registry = default_registry()

    assert registry.get("HK-HKD").adapter_map["hkma_official"].retry_limit == 0
    assert registry.get("NZ-NZD").adapter_map["rbnz_policy"].retry_limit == 0
    assert registry.get("NZ-NZD").adapter_map["rbnz_wholesale"].retry_limit == 0


@pytest.mark.asyncio
async def test_hkma_retries_server_errors_with_bounded_backoff(
    monkeypatch,
) -> None:
    statuses = iter((502, 503, 200))
    requests: list[httpx.Request] = []
    delays: list[float] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        status = next(statuses)
        return httpx.Response(
            status,
            headers={"content-type": "application/json; charset=utf-8"},
            json={"result": {"records": []}},
        )

    async def record_sleep(delay: float) -> None:
        delays.append(delay)

    monkeypatch.setattr(official, "_sleep", record_sleep)
    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        documents = tuple(await _official_adapter("hkma_official").fetcher(client))

    assert [request.url.params["pagesize"] for request in requests] == [
        "1000",
        "1000",
        "1000",
    ]
    assert all(request.headers["accept"] == "application/json" for request in requests)
    assert delays == [1.0, 2.0]
    assert len(documents) == 1
    assert documents[0].label == "hkma_liquidity"
    assert documents[0].media_type == "application/json"
    assert documents[0].source_uri == str(requests[-1].url)


@pytest.mark.asyncio
async def test_hkma_retries_request_errors_within_the_same_budget(
    monkeypatch,
) -> None:
    requests: list[httpx.Request] = []
    delays: list[float] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        if len(requests) < 3:
            raise httpx.ConnectError("temporary connect failure", request=request)
        return httpx.Response(200, json={"result": {"records": []}})

    async def record_sleep(delay: float) -> None:
        delays.append(delay)

    monkeypatch.setattr(official, "_sleep", record_sleep)
    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        documents = tuple(await _official_adapter("hkma_official").fetcher(client))

    assert len(requests) == 3
    assert delays == [1.0, 2.0]
    assert documents[0].source_uri == str(requests[-1].url)


@pytest.mark.asyncio
async def test_hkma_raises_last_server_response_after_retry_exhaustion(
    monkeypatch,
) -> None:
    requests: list[httpx.Request] = []
    delays: list[float] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        return httpx.Response(502, text="upstream temporarily unavailable")

    async def record_sleep(delay: float) -> None:
        delays.append(delay)

    monkeypatch.setattr(official, "_sleep", record_sleep)
    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        with pytest.raises(httpx.HTTPStatusError) as raised:
            await _official_adapter("hkma_official").fetcher(client)

    assert raised.value.response.status_code == 502
    assert len(requests) == 3
    assert delays == [1.0, 2.0]


@pytest.mark.asyncio
async def test_hkma_does_not_retry_a_non_server_response(monkeypatch) -> None:
    requests: list[httpx.Request] = []
    delays: list[float] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        return httpx.Response(403, text="forbidden")

    async def record_sleep(delay: float) -> None:
        delays.append(delay)

    monkeypatch.setattr(official, "_sleep", record_sleep)
    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        with pytest.raises(httpx.HTTPStatusError):
            await _official_adapter("hkma_official").fetcher(client)

    assert len(requests) == 1
    assert delays == []


@pytest.mark.asyncio
async def test_rbnz_uses_canonical_official_html_after_workbook_403(
    monkeypatch,
) -> None:
    _approve_rbnz_access(monkeypatch)
    page = b"""
        <html><body><table>
          <tr><th>Cash rate (%pa)</th></tr>
          <tr>
            <th>Date</th>
            <th>Official Cash Rate (OCR)</th>
            <th>Overnight Deposit Rate</th>
            <th>Overnight Reverse Repurchase Facility Rate</th>
          </tr>
          <tr><td>08 Aug 2026</td><td>3.25</td><td>3.20</td><td>3.50</td></tr>
        </table>
        <table>
          <tr><td>09 Aug 2026</td><td>9.25</td><td>9.20</td><td>9.50</td></tr>
        </table>
        </body></html>
    """

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith(".xlsx"):
            return httpx.Response(403, text="forbidden")
        return httpx.Response(
            200,
            headers={"content-type": "text/html; charset=utf-8"},
            content=page,
        )

    transport, requests = _rbnz_mock_transport(handler)
    async with httpx.AsyncClient(transport=transport) as client:
        documents = tuple(await _official_adapter("rbnz_wholesale").fetcher(client))

    assert [str(request.url) for request in requests] == [
        "https://www.rbnz.govt.nz/-/media/project/sites/rbnz/files/statistics/"
        "series/b/b2/hb2-daily-close.xlsx",
        "https://www.rbnz.govt.nz/en/statistics/series/"
        "exchange-and-interest-rates/wholesale-interest-rates",
    ]
    assert (
        requests[0]
        .headers["accept"]
        .startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    )
    assert requests[0].headers["referer"] == str(requests[1].url)
    assert all(
        request.headers["user-agent"] == official.USER_AGENT for request in requests
    )
    assert all("sec-ch-ua" not in request.headers for request in requests)
    assert documents[0].source_uri == str(requests[1].url)
    assert documents[0].media_type == "text/html"
    assert {
        point.instrument_id: point.raw_value for point in parse_rbnz(documents[0])
    } == {
        "NZ.RBNZ.OVERNIGHT_DEPOSIT": Decimal("3.20"),
        "NZ.RBNZ.OVERNIGHT_REVERSE_REPO": Decimal("3.50"),
    }


@pytest.mark.asyncio
async def test_rbnz_uses_the_official_workbook_without_requesting_fallback(
    monkeypatch,
) -> None:
    _approve_rbnz_access(monkeypatch)

    def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(
            200,
            headers={
                "content-type": (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            },
            content=_rbnz_workbook_payload(),
        )

    transport, requests = _rbnz_mock_transport(handler)
    async with httpx.AsyncClient(transport=transport) as client:
        documents = tuple(await _official_adapter("rbnz_policy").fetcher(client))

    assert len(requests) == 1
    assert len(documents) == 1
    assert documents[0].source_uri == str(requests[0].url)
    assert documents[0].source_uri.endswith("hb2-daily-close.xlsx")
    assert documents[0].media_type.endswith("spreadsheetml.sheet")
    assert documents[0].payload.startswith(b"PK\x03\x04")


@pytest.mark.asyncio
async def test_rbnz_never_follows_a_redirect(
    monkeypatch,
) -> None:
    _approve_rbnz_access(monkeypatch)

    def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(
            302,
            headers={"location": "https://mirror.example.invalid/rbnz-b2.xlsx"},
        )

    transport, requests = _rbnz_mock_transport(handler)
    async with httpx.AsyncClient(transport=transport) as client:
        with pytest.raises(RBNZSourceUnavailableError) as raised:
            await _official_adapter("rbnz_policy").fetcher(client)

    assert len(requests) == 2
    assert str(raised.value).count("HTTP 302") == 2
    assert all(request.url.host == "www.rbnz.govt.nz" for request in requests)


def test_rbnz_parses_the_current_series_id_workbook_contract() -> None:
    document = FetchedDocument(
        official._RBNZ_B2_XLSX_URI,
        official._RBNZ_XLSX_MEDIA_TYPE,
        _rbnz_current_workbook_payload(),
        "rbnz_wholesale",
    )

    points = {point.instrument_id: point for point in parse_rbnz(document)}

    assert points["NZ.RBNZ.OVERNIGHT_DEPOSIT"].raw_value == Decimal("2.5")
    assert points["NZ.RBNZ.OVERNIGHT_REVERSE_REPO"].raw_value == Decimal("3")


def test_rbnz_modern_workbook_cannot_downgrade_to_display_headings() -> None:
    document = FetchedDocument(
        official._RBNZ_B2_XLSX_URI,
        official._RBNZ_XLSX_MEDIA_TYPE,
        _rbnz_current_workbook_payload(reverse_series_id="CHANGED.UPSTREAM.ID"),
        "rbnz_wholesale",
    )

    with pytest.raises(ValueError, match="exactly one usable B2 data sheet"):
        parse_rbnz(document)


def test_rbnz_renamed_series_marker_cannot_downgrade_to_display_headings() -> None:
    document = FetchedDocument(
        official._RBNZ_B2_XLSX_URI,
        official._RBNZ_XLSX_MEDIA_TYPE,
        _rbnz_current_workbook_payload(series_marker="Series identifier"),
        "rbnz_wholesale",
    )

    with pytest.raises(ValueError, match="exactly one usable B2 data sheet"):
        parse_rbnz(document)


def test_rbnz_modern_signals_reject_a_separate_legacy_candidate() -> None:
    document = FetchedDocument(
        official._RBNZ_B2_XLSX_URI,
        official._RBNZ_XLSX_MEDIA_TYPE,
        _rbnz_current_workbook_payload(
            series_marker="Identifier",
            add_legacy_candidate=True,
        ),
        "rbnz_wholesale",
    )

    with pytest.raises(ValueError, match="exactly one usable B2 data sheet"):
        parse_rbnz(document)


def test_rbnz_modern_workbook_rejects_duplicate_series_columns() -> None:
    document = FetchedDocument(
        official._RBNZ_B2_XLSX_URI,
        official._RBNZ_XLSX_MEDIA_TYPE,
        _rbnz_current_workbook_payload(duplicate_reverse=True),
        "rbnz_wholesale",
    )

    with pytest.raises(ValueError, match="exactly one usable B2 data sheet"):
        parse_rbnz(document)


def test_rbnz_workbook_archive_row_and_cell_bounds_fail_closed(monkeypatch) -> None:
    document = FetchedDocument(
        official._RBNZ_B2_XLSX_URI,
        official._RBNZ_XLSX_MEDIA_TYPE,
        _rbnz_current_workbook_payload(),
        "rbnz_wholesale",
    )

    monkeypatch.setattr(official, "_RBNZ_MAX_XLSX_EXPANDED_BYTES", 1)
    with pytest.raises(ValueError, match="XLSX expansion limit"):
        parse_rbnz(document)
    monkeypatch.setattr(official, "_RBNZ_MAX_XLSX_EXPANDED_BYTES", 64 * 1024 * 1024)
    monkeypatch.setattr(official, "_RBNZ_MAX_WORKBOOK_ROWS", 1)
    with pytest.raises(ValueError, match="row limit"):
        parse_rbnz(document)
    monkeypatch.setattr(official, "_RBNZ_MAX_WORKBOOK_ROWS", 50_000)
    monkeypatch.setattr(official, "_RBNZ_MAX_WORKBOOK_CELLS", 1)
    with pytest.raises(ValueError, match="cell limit"):
        parse_rbnz(document)


@pytest.mark.asyncio
async def test_cfets_access_defaults_off_before_any_request(monkeypatch) -> None:
    monkeypatch.delenv(official._CFETS_APPROVAL_PATH_ENV, raising=False)
    monkeypatch.delenv(official._CFETS_APPROVAL_SHA256_ENV, raising=False)
    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        return httpx.Response(500, request=request)

    adapter = _official_adapter("cfets_rates")
    with pytest.raises(SourcePolicyUnavailableError) as availability:
        adapter.check_availability()
    assert official._CFETS_APPROVAL_PATH_ENV in str(availability.value)
    assert official._CFETS_APPROVAL_SHA256_ENV in str(availability.value)

    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        with pytest.raises(SourcePolicyUnavailableError, match="operator-held"):
            await adapter.fetcher(client)

    assert requests == []


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("failure", "message"),
    (
        ("missing", "missing or cannot be opened safely"),
        ("symlink", "missing or cannot be opened safely"),
        ("mode", "ownership, mode, or file type is unsafe"),
        ("owner", "ownership, mode, or file type is unsafe"),
        ("schema", "missing or unknown fields"),
        ("hash", "digest does not match"),
        ("path", "operator-held data licence"),
        ("scope_endpoint", "does not grant the exact endpoints"),
        ("scope_product", "does not grant the exact endpoints"),
        ("scope_output", "does not grant the exact endpoints"),
        ("scope_use", "does not grant the exact endpoints"),
        ("scope_publication", "does not grant the exact endpoints"),
        ("scope_retention", "does not grant the exact endpoints"),
        ("evidence_missing", "licence evidence is missing"),
        ("evidence_symlink", "licence evidence is missing"),
        ("evidence_mode", "licence evidence ownership, mode"),
        ("evidence_hash", "licence evidence digest does not match"),
        ("evidence_path", "does not name the canonical licence evidence"),
    ),
)
async def test_cfets_unsafe_or_unscoped_artifact_never_reaches_network(
    tmp_path,
    monkeypatch,
    failure: str,
    message: str,
) -> None:
    kwargs = {}
    if failure == "schema":
        kwargs["extra_fields"] = {"unexpected": "field"}
    elif failure == "scope_endpoint":
        kwargs["overrides"] = {"endpoints": official._CFETS_SHIBOR_ON_ENDPOINT}
    elif failure == "scope_product":
        kwargs["overrides"] = {"upstream_products": "SHIBOR_ON"}
    elif failure == "scope_output":
        kwargs["overrides"] = {"canonical_outputs": "CN.CFETS.SHIBOR_ON"}
    elif failure == "scope_use":
        kwargs["overrides"] = {"permitted_use": "commercial_analytics"}
    elif failure == "scope_publication":
        kwargs["overrides"] = {"publication": "allowed"}
    elif failure == "scope_retention":
        kwargs["overrides"] = {"raw_response_retention": "allowed"}
    elif failure == "evidence_hash":
        kwargs["overrides"] = {"licence_evidence_sha256": "d" * 64}
    elif failure == "evidence_path":
        kwargs["overrides"] = {
            "licence_evidence_path": str(tmp_path / "redirected-evidence.pdf")
        }
    approval_path, _ = _approve_cfets_access(tmp_path, monkeypatch, **kwargs)
    evidence_path = official._CFETS_LICENCE_EVIDENCE_CANONICAL_PATH
    if failure == "missing":
        approval_path.unlink()
    elif failure == "symlink":
        payload = approval_path.read_bytes()
        target = tmp_path / "approval-target.conf"
        target.write_bytes(payload)
        target.chmod(0o640)
        approval_path.unlink()
        approval_path.symlink_to(target)
    elif failure == "mode":
        approval_path.chmod(0o644)
    elif failure == "owner":
        monkeypatch.setattr(
            official,
            "_cfets_approval_expected_owner",
            lambda: (os.getuid() + 1, os.getgid()),
        )
    elif failure == "hash":
        monkeypatch.setenv(official._CFETS_APPROVAL_SHA256_ENV, "d" * 64)
    elif failure == "path":
        monkeypatch.setenv(
            official._CFETS_APPROVAL_PATH_ENV,
            str(tmp_path / "redirected-approval.conf"),
        )
    elif failure == "evidence_missing":
        evidence_path.unlink()
    elif failure == "evidence_symlink":
        payload = evidence_path.read_bytes()
        target = tmp_path / "evidence-target.pdf"
        target.write_bytes(payload)
        target.chmod(0o640)
        evidence_path.unlink()
        evidence_path.symlink_to(target)
    elif failure == "evidence_mode":
        evidence_path.chmod(0o644)

    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        return httpx.Response(500, request=request)

    adapter = _official_adapter("cfets_rates")
    with pytest.raises(SourcePolicyUnavailableError, match=message):
        adapter.check_availability()
    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        with pytest.raises(SourcePolicyUnavailableError, match=message):
            await adapter.fetcher(client)

    assert requests == []


@pytest.mark.asyncio
async def test_cfets_valid_bounded_licence_proof_enables_collection(
    tmp_path,
    monkeypatch,
) -> None:
    _, artifact_sha256 = _approve_cfets_access(tmp_path, monkeypatch)
    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        return _cfets_success_response(request)

    adapter = _official_adapter("cfets_rates")
    adapter.check_availability()
    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        documents = tuple(await adapter.fetcher(client))

    assert [document.label for document in documents] == [
        "CN.CFETS.FDR007",
        "CN.CFETS.SHIBOR_ON",
    ]
    assert [request.url.path for request in requests] == [
        "/r/cms/www/chinamoney/data/currency/fdr-settings.json",
        "/r/cms/www/chinamoney/data/currency/fdr-chrt.csv",
        "/ags/ms/cm-u-bk-shibor/ShiborHis",
    ]
    generation = f"cfets-approval-v2-{artifact_sha256[:16]}"
    assert all(
        urllib.parse.parse_qs(urllib.parse.urlsplit(document.source_uri).fragment)
        == {official._CFETS_RIGHTS_GENERATION_KEY: [generation]}
        for document in documents
    )
    points = tuple(
        point
        for document in documents
        for point in official.parse_cfets_rates(document)
    )
    assert all(
        re.fullmatch(
            rf"cfets:{generation}:2026-08-11:[0-9a-f]{{16}}",
            str(point.revision_id),
        )
        for point in points
    )
    retained = [json.loads(document.payload) for document in documents]
    assert all(
        projection["raw_response_retained"] is False
        and projection["columns"] == ["event_date", "value"]
        and set(projection["records"][0]) == {"event_date", "value"}
        for projection in retained
    )
    assert all(b"FDR001" not in document.payload for document in documents)
    assert all(b"FDR014" not in document.payload for document in documents)


@pytest.mark.asyncio
async def test_cfets_single_window_final_response_revocation_persists_nothing(
    tmp_path,
    monkeypatch,
) -> None:
    _approve_cfets_access(tmp_path, monkeypatch)
    evidence_path = official._CFETS_LICENCE_EVIDENCE_CANONICAL_PATH
    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        response = _cfets_success_response(request)
        if request.url.path.endswith("/ShiborHis"):
            evidence_path.unlink()
        return response

    probe, runs, raw_sink, normalized_sink, observation_writes = (
        await _run_cfets_final_response_probe(
            _official_adapter("cfets_rates"),
            handler,
        )
    )

    assert [request.url.path for request in requests] == [
        "/r/cms/www/chinamoney/data/currency/fdr-settings.json",
        "/r/cms/www/chinamoney/data/currency/fdr-chrt.csv",
        "/ags/ms/cm-u-bk-shibor/ShiborHis",
    ]
    assert probe.fetch_completed is False
    assert runs[0].status is CollectorRunStatus.UNAVAILABLE
    assert raw_sink.writes == []
    assert normalized_sink.writes == []
    assert observation_writes == []


@pytest.mark.asyncio
async def test_cfets_single_window_midnight_expiry_persists_nothing(
    tmp_path,
    monkeypatch,
) -> None:
    _approve_cfets_access(
        tmp_path,
        monkeypatch,
        valid_until="2026-08-14",
    )
    current_day = [date(2026, 8, 14)]
    monkeypatch.setattr(official, "_cfets_access_today", lambda: current_day[0])
    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        response = _cfets_success_response(request)
        if request.url.path.endswith("/ShiborHis"):
            current_day[0] = date(2026, 8, 15)
        return response

    probe, runs, raw_sink, normalized_sink, observation_writes = (
        await _run_cfets_final_response_probe(
            _official_adapter("cfets_rates"),
            handler,
        )
    )

    assert [request.url.path for request in requests] == [
        "/r/cms/www/chinamoney/data/currency/fdr-settings.json",
        "/r/cms/www/chinamoney/data/currency/fdr-chrt.csv",
        "/ags/ms/cm-u-bk-shibor/ShiborHis",
    ]
    assert probe.fetch_completed is False
    assert runs[0].status is CollectorRunStatus.UNAVAILABLE
    assert raw_sink.writes == []
    assert normalized_sink.writes == []
    assert observation_writes == []


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("corruption", "message"),
    (
        ("schema_columns", "chart columns changed"),
        ("fdr_width", "does not match named columns"),
        ("shibor_fields", "changed named fields"),
        ("shibor_window", "outside the approved query window"),
    ),
)
async def test_cfets_named_schema_changes_fail_before_retained_documents(
    tmp_path,
    monkeypatch,
    corruption: str,
    message: str,
) -> None:
    _approve_cfets_access(tmp_path, monkeypatch)
    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        if corruption == "schema_columns" and request.url.path.endswith(
            "/fdr-settings.json"
        ):
            response = _cfets_success_response(request)
            payload = response.json()
            payload["columns"][7] = "DR007"
            return httpx.Response(200, request=request, json=payload)
        if corruption == "fdr_width" and request.url.path.endswith("/fdr-chrt.csv"):
            return httpx.Response(200, request=request, content=b"2026-08-11,1.52\n")
        if corruption == "shibor_fields" and request.url.path.endswith("/ShiborHis"):
            return httpx.Response(
                200,
                request=request,
                json={"records": [{"showDateCN": "2026-08-11", "1W": "1.31"}]},
            )
        if corruption == "shibor_window" and request.url.path.endswith("/ShiborHis"):
            return httpx.Response(
                200,
                request=request,
                json={"records": [{"showDateCN": "1999-01-01", "ON": "1.31"}]},
            )
        return _cfets_success_response(request)

    adapter = _official_adapter("cfets_rates")
    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        with pytest.raises(ValueError, match=message):
            await adapter.fetcher(client)

    assert requests


@pytest.mark.asyncio
async def test_cfets_exact_endpoint_scope_does_not_follow_redirects(
    tmp_path,
    monkeypatch,
) -> None:
    _approve_cfets_access(tmp_path, monkeypatch)
    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        return httpx.Response(
            302,
            request=request,
            headers={"Location": "https://example.invalid/unapproved-mirror"},
        )

    adapter = _official_adapter("cfets_rates")
    async with httpx.AsyncClient(
        transport=httpx.MockTransport(handler),
        follow_redirects=True,
    ) as client:
        with pytest.raises(httpx.HTTPStatusError):
            await adapter.fetcher(client)

    assert [str(request.url) for request in requests] == [
        official._CFETS_SCHEMA_ENDPOINT
    ]


def test_cfets_parser_refuses_unminimized_upstream_bodies() -> None:
    raw = FetchedDocument(
        official._CFETS_SHIBOR_ON_ENDPOINT,
        "application/json",
        b'{"records":[{"showDateCN":"2026-08-11","ON":"1.31",'
        b'"1W":"1.40"}]}',
        "CN.CFETS.SHIBOR_ON",
    )

    with pytest.raises(ValueError, match="retained projection"):
        official.parse_cfets_rates(raw)


def test_cfets_source_catalog_remains_metadata_only() -> None:
    spec = default_registry().get("CN-CNY").adapter_map["cfets_rates"]
    instruments = default_registry().get("CN-CNY").instrument_map

    assert spec.classification is ConnectorClassification.LICENSED
    assert spec.redistribution_status is RedistributionStatus.METADATA_ONLY
    assert ("CN-CNY", "cfets_rates") in PRODUCTION_ADAPTER_KEYS
    assert "CN.CFETS.FDR007" in instruments
    assert "CN.CFETS.DR007" not in instruments


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("valid_until", "message"),
    (
        ("2026-08-13", "review has expired"),
        ("2027-08-16", "reviewed within 366 days"),
    ),
)
async def test_cfets_expired_or_overlong_proof_stays_offline(
    tmp_path,
    monkeypatch,
    valid_until: str,
    message: str,
) -> None:
    _approve_cfets_access(tmp_path, monkeypatch, valid_until=valid_until)
    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        return httpx.Response(500, request=request)

    adapter = _official_adapter("cfets_rates")
    with pytest.raises(SourcePolicyUnavailableError, match=message):
        adapter.check_availability()
    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        with pytest.raises(SourcePolicyUnavailableError, match=message):
            await adapter.fetcher(client)

    assert requests == []


@pytest.mark.asyncio
async def test_cfets_reuses_adapter_with_a_fresh_date_for_each_collection(
    tmp_path,
    monkeypatch,
) -> None:
    _approve_cfets_access(tmp_path, monkeypatch)
    dates = iter(
        (date(2026, 8, 14),) * 7
        + (date(2026, 8, 15),) * 7
    )
    monkeypatch.setattr(official, "_cfets_access_today", lambda: next(dates))
    shibor_end_dates: list[str] = []

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith("/ShiborHis"):
            shibor_end_dates.append(request.url.params["endDate"])
        return _cfets_success_response(request)

    adapter = _official_adapter("cfets_rates")
    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        await adapter.fetcher(client)
        await adapter.fetcher(client)

    assert shibor_end_dates == ["2026-08-14", "2026-08-15"]


@pytest.mark.asyncio
async def test_cfets_expiry_between_backfill_chunks_stops_before_next_request(
    tmp_path,
    monkeypatch,
) -> None:
    _approve_cfets_access(tmp_path, monkeypatch, valid_until="2026-08-14")
    dates = iter((date(2026, 8, 14),) * 7 + (date(2026, 8, 15),))
    monkeypatch.setattr(official, "_cfets_access_today", lambda: next(dates))
    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        return _cfets_success_response(request)

    adapter = _official_adapter("cfets_rates", backfill=True)
    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        with pytest.raises(SourcePolicyUnavailableError, match="review has expired"):
            await adapter.fetcher(client)

    assert [request.url.path for request in requests] == [
        "/r/cms/www/chinamoney/data/currency/fdr-settings.json",
        "/r/cms/www/chinamoney/data/currency/fdr-chrt.csv",
        "/ags/ms/cm-u-bk-shibor/ShiborHis",
    ]


@pytest.mark.asyncio
async def test_cfets_revocation_between_backfill_chunks_stops_next_request(
    tmp_path,
    monkeypatch,
) -> None:
    approval_path, _ = _approve_cfets_access(tmp_path, monkeypatch)
    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        response = _cfets_success_response(request)
        if request.url.path.endswith("/ShiborHis"):
            approval_path.unlink()
        return response

    adapter = _official_adapter("cfets_rates", backfill=True)
    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        with pytest.raises(SourcePolicyUnavailableError, match="missing"):
            await adapter.fetcher(client)

    assert [request.url.path for request in requests] == [
        "/r/cms/www/chinamoney/data/currency/fdr-settings.json",
        "/r/cms/www/chinamoney/data/currency/fdr-chrt.csv",
        "/ags/ms/cm-u-bk-shibor/ShiborHis",
    ]


@pytest.mark.asyncio
async def test_rbnz_access_defaults_off_without_written_approval(monkeypatch) -> None:
    monkeypatch.delenv(official._RBNZ_ACCESS_APPROVAL_SHA256_ENV, raising=False)
    monkeypatch.delenv(
        official._RBNZ_ACCESS_APPROVAL_VALID_UNTIL_ENV,
        raising=False,
    )
    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        return httpx.Response(200, content=_rbnz_workbook_payload())

    adapter = _official_adapter("rbnz_policy")
    with pytest.raises(SourcePolicyUnavailableError, match="prior written permission"):
        adapter.check_availability()

    async with httpx.AsyncClient(transport=httpx.MockTransport(handler)) as client:
        with pytest.raises(
            RBNZSourceUnavailableError, match="prior written permission"
        ):
            await adapter.fetcher(client)

    assert requests == []


@pytest.mark.parametrize(
    ("valid_until", "message"),
    (
        ("2026-08-13", "review has expired"),
        ("2027-08-16", "reviewed within 366 days"),
    ),
)
def test_rbnz_access_approval_has_a_bounded_review_window(
    monkeypatch,
    valid_until: str,
    message: str,
) -> None:
    monkeypatch.setattr(official, "_rbnz_access_today", lambda: date(2026, 8, 14))
    monkeypatch.setenv(official._RBNZ_ACCESS_APPROVAL_SHA256_ENV, "a" * 64)
    monkeypatch.setenv(
        official._RBNZ_ACCESS_APPROVAL_VALID_UNTIL_ENV,
        valid_until,
    )

    with pytest.raises(RBNZSourceUnavailableError, match=message):
        official._require_rbnz_access_approval()


@pytest.mark.asyncio
async def test_rbnz_transport_bounds_decoded_response_body(monkeypatch) -> None:
    _approve_rbnz_access(monkeypatch)
    monkeypatch.setattr(official, "_RBNZ_MAX_BODY_BYTES", 4)

    def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(200, content=b"12345")

    transport, requests = _rbnz_mock_transport(handler)
    async with httpx.AsyncClient(transport=transport) as client:
        with pytest.raises(RBNZSourceUnavailableError) as raised:
            await _official_adapter("rbnz_policy").fetcher(client)

    assert len(requests) == 2
    assert str(raised.value).count("body limit") == 2


@pytest.mark.asyncio
async def test_rbnz_transport_rejects_compression_before_decoding(monkeypatch) -> None:
    _approve_rbnz_access(monkeypatch)
    compressed = gzip.compress(b"x" * (official._RBNZ_MAX_BODY_BYTES + 1))

    def handler(request: httpx.Request) -> httpx.Response:
        assert request.headers["accept-encoding"] == "identity"
        return httpx.Response(
            200,
            headers={"content-encoding": "gzip"},
            stream=httpx.ByteStream(compressed),
        )

    transport, requests = _rbnz_mock_transport(handler)
    async with httpx.AsyncClient(transport=transport) as client:
        with pytest.raises(RBNZSourceUnavailableError) as raised:
            await _official_adapter("rbnz_policy").fetcher(client)

    assert len(requests) == 2
    assert str(raised.value).count("unsupported transport content encoding") == 2


@pytest.mark.asyncio
async def test_rbnz_transport_enforces_a_total_response_deadline(monkeypatch) -> None:
    _approve_rbnz_access(monkeypatch)
    monkeypatch.setattr(official, "_RBNZ_TOTAL_RESPONSE_TIMEOUT_SECONDS", 0.01)

    class SlowStream(httpx.AsyncByteStream):
        async def __aiter__(self):
            await asyncio.sleep(0.05)
            yield b"PK\x03\x04"

    def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(200, stream=SlowStream())

    transport, requests = _rbnz_mock_transport(handler)
    async with httpx.AsyncClient(transport=transport) as client:
        with pytest.raises(RBNZSourceUnavailableError) as raised:
            await _official_adapter("rbnz_policy").fetcher(client)

    assert len(requests) == 2
    assert str(raised.value).count("total response deadline") == 2


@pytest.mark.asyncio
async def test_rbnz_unusable_pk_workbook_gets_one_canonical_html_attempt(
    monkeypatch,
) -> None:
    _approve_rbnz_access(monkeypatch)
    page = b"""
        <table>
          <tr>
            <th>Date</th>
            <th>Official Cash Rate (OCR)</th>
            <th>Overnight Deposit Rate</th>
            <th>Overnight Reverse Repurchase Facility Rate</th>
          </tr>
          <tr><td>10 Aug 2026</td><td>2.50</td><td>2.50</td><td>3.00</td></tr>
        </table>
    """

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith(".xlsx"):
            return httpx.Response(
                200,
                headers={"content-type": "application/octet-stream"},
                content=_rbnz_workbook_payload(include_b2=False),
            )
        return httpx.Response(200, headers={"content-type": "text/html"}, content=page)

    transport, requests = _rbnz_mock_transport(handler)
    async with httpx.AsyncClient(transport=transport) as client:
        documents = tuple(await _official_adapter("rbnz_policy").fetcher(client))

    assert len(requests) == 2
    assert requests[0].url.path.endswith("hb2-daily-close.xlsx")
    assert requests[1].url.path.endswith("wholesale-interest-rates")
    assert documents[0].source_uri == str(requests[1].url)
    assert documents[0].media_type == "text/html"


def test_rbnz_html_evidence_is_scoped_to_each_instrument_cell() -> None:
    def parse(*, ocr: str, deposit: str, reverse_repo: str):
        document = FetchedDocument(
            "https://www.rbnz.govt.nz/statistics/series/exchange-and-interest-rates/"
            "wholesale-interest-rates",
            "text/html",
            f"""
                <table>
                  <tr><th>Date</th><th>Official Cash Rate (OCR)</th>
                    <th>Overnight Deposit Rate</th>
                    <th>Overnight Reverse Repurchase Facility Rate</th></tr>
                  <tr><td>10 Aug 2026</td><td>{ocr}</td><td>{deposit}</td>
                    <td>{reverse_repo}</td></tr>
                </table>
            """.encode(),
            "rbnz_wholesale",
        )
        return {point.instrument_id: point for point in parse_rbnz(document)}

    first = parse(ocr="2.50", deposit="2.50", reverse_repo="3.00")
    changed = parse(ocr="9.99", deposit="2.50", reverse_repo="3.10")

    assert (
        first["NZ.RBNZ.OVERNIGHT_DEPOSIT"].row_evidence
        == changed["NZ.RBNZ.OVERNIGHT_DEPOSIT"].row_evidence
    )
    assert (
        first["NZ.RBNZ.OVERNIGHT_REVERSE_REPO"].row_evidence
        != changed["NZ.RBNZ.OVERNIGHT_REVERSE_REPO"].row_evidence
    )


@pytest.mark.asyncio
async def test_rbnz_403_exhaustion_reports_both_official_endpoints(monkeypatch) -> None:
    _approve_rbnz_access(monkeypatch)

    def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(403, text="cloudflare challenge")

    transport, requests = _rbnz_mock_transport(handler)
    async with httpx.AsyncClient(transport=transport) as client:
        with pytest.raises(RBNZSourceUnavailableError) as raised:
            await _official_adapter("rbnz_policy").fetcher(client)

    detail = str(raised.value)
    assert len(requests) == 2
    assert detail.count("HTTP 403") == 2
    assert "hb2-daily-close.xlsx" in detail
    assert "/statistics/series/exchange-and-interest-rates/" in detail


@pytest.mark.asyncio
async def test_rbnz_rejects_html_without_the_expected_b2_table(monkeypatch) -> None:
    _approve_rbnz_access(monkeypatch)

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith(".xlsx"):
            return httpx.Response(403, text="forbidden")
        return httpx.Response(
            200,
            headers={"content-type": "text/html"},
            text="<html><table><tr><th>Date</th><th>Unrelated</th></tr>"
            "<tr><td>08 Aug 2026</td><td>9.99</td></tr></table></html>",
        )

    transport, _ = _rbnz_mock_transport(handler)
    async with httpx.AsyncClient(transport=transport) as client:
        with pytest.raises(RBNZSourceUnavailableError) as raised:
            await _official_adapter("rbnz_policy").fetcher(client)

    detail = str(raised.value)
    assert "HTTP 403" in detail
    assert "no expected B2 interest-rate table header" in detail


@pytest.mark.asyncio
async def test_rbnz_backfill_rejects_the_bounded_html_summary(monkeypatch) -> None:
    _approve_rbnz_access(monkeypatch)
    page = b"""
        <table>
          <tr>
            <th>Date</th>
            <th>Official Cash Rate (OCR)</th>
            <th>Overnight Deposit Rate</th>
            <th>Overnight Reverse Repurchase Facility Rate</th>
          </tr>
          <tr><td>10 Aug 2026</td><td>2.50</td><td>2.50</td><td>3.00</td></tr>
        </table>
    """

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith(".xlsx"):
            return httpx.Response(403, text="forbidden")
        return httpx.Response(200, headers={"content-type": "text/html"}, content=page)

    transport, _ = _rbnz_mock_transport(handler)
    async with httpx.AsyncClient(transport=transport) as client:
        with pytest.raises(RBNZSourceUnavailableError) as raised:
            await _official_adapter("rbnz_policy", backfill=True).fetcher(client)

    detail = str(raised.value)
    assert "HTTP 403" in detail
    assert "recent summary" in detail
    assert "cannot satisfy a historical backfill" in detail


@pytest.mark.asyncio
async def test_historical_current_vintage_is_not_leaked_into_the_past(
    tmp_path, monkeypatch
) -> None:
    monkeypatch.setattr(store, "DB_PATH", tmp_path / "backfill.sqlite")
    capture = datetime(2026, 8, 9, 12, tzinfo=UTC)
    document = FetchedDocument(
        "https://example.invalid/official.csv",
        "text/csv",
        b"DATE,VALUE\n2020-01-02,1.50\n",
        "US.NYFED.SOFR",
    )

    async def fetcher(_client):
        return (document,)

    def parser(_document):
        return (
            ParsedPoint(
                "US.NYFED.SOFR",
                date(2020, 1, 2),
                "1.50",
                b"2020-01-02,1.50",
            ),
        )

    adapter = FunctionalCanonicalAdapter(
        pack=default_registry().get("US-USD"),
        adapter_id="fred_daily",
        source="official-test",
        fetcher=fetcher,
        parser=parser,
        repository=SQLiteMarketRepository(),
        clock=lambda: capture,
        historical_backfill=True,
    )
    batch = await adapter.collect()
    observation = batch.observations[0]

    assert observation.source_publication_time < capture
    assert observation.knowledge_time == capture
    assert observation.quality is QualityState.PROVISIONAL
    assert observation.value == 150


@pytest.mark.asyncio
async def test_changed_source_content_can_revert_without_revision_id_collision(
    tmp_path, monkeypatch
) -> None:
    monkeypatch.setattr(store, "DB_PATH", tmp_path / "revision-reversion.sqlite")
    repository = SQLiteMarketRepository()
    base = (
        b'{"refRates":[{"type":"SOFR","effectiveDate":"2026-08-06",'
        b'"percentRate":5.30,"percentPercentile99":5.40,'
        b'"volumeInBillions":2000}]}'
    )
    changed = base.replace(b'"percentRate":5.30', b'"percentRate":5.31')

    async def collect(payload: bytes, captured_at: datetime):
        document = FetchedDocument(
            "https://markets.newyorkfed.org/api/rates/secured/all/search.json",
            "application/json",
            payload,
            "nyfed_secured_rates",
        )

        async def fetcher(_client):
            return (document,)

        adapter = FunctionalCanonicalAdapter(
            pack=default_registry().get("US-USD"),
            adapter_id="nyfed_rates",
            source="nyfed_rates",
            fetcher=fetcher,
            parser=parse_nyfed_rates,
            repository=repository,
            clock=lambda: captured_at,
        )
        batch = await adapter.collect()
        repository.save_observations(batch.observations)
        return next(
            row
            for row in batch.observations
            if row.instrument_id == "US.NYFED.SOFR_MEDIAN"
        )

    first = await collect(base, datetime(2026, 8, 10, 10, tzinfo=UTC))
    second = await collect(changed, datetime(2026, 8, 10, 11, tzinfo=UTC))
    reverted = await collect(base, datetime(2026, 8, 10, 12, tzinfo=UTC))

    assert len({first.revision_id, second.revision_id, reverted.revision_id}) == 3
    assert "@capture-20260810T110000Z" in second.revision_id
    assert "@capture-20260810T120000Z" in reverted.revision_id


@pytest.mark.asyncio
async def test_identical_content_gains_explicit_field_lineage_once(
    tmp_path, monkeypatch
) -> None:
    monkeypatch.setattr(store, "DB_PATH", tmp_path / "lineage-upgrade.sqlite")
    repository = SQLiteMarketRepository()
    document = FetchedDocument(
        "https://markets.newyorkfed.org/api/rates/secured/all/search.json",
        "application/json",
        b"{}",
        "nyfed_secured_rates",
    )
    evidence = b"same canonical P99 source row"
    explicit_revision = "nyfed:percentPercentile99:2026-08-06:unrevised-content"

    async def collect(*, captured_at: datetime, revision_id: str | None):
        async def fetcher(_client):
            return (document,)

        def parser(_document):
            return (
                ParsedPoint(
                    "US.NYFED.SOFR_P99",
                    date(2026, 8, 6),
                    Decimal("5.40"),
                    evidence,
                    revision_id=revision_id,
                ),
            )

        adapter = FunctionalCanonicalAdapter(
            pack=default_registry().get("US-USD"),
            adapter_id="nyfed_rates",
            source="nyfed_rates",
            fetcher=fetcher,
            parser=parser,
            repository=repository,
            clock=lambda: captured_at,
        )
        batch = await adapter.collect()
        repository.save_observations(batch.observations)
        return batch.observations[0]

    legacy = await collect(
        captured_at=datetime(2026, 8, 10, 10, tzinfo=UTC),
        revision_id=None,
    )
    upgraded = await collect(
        captured_at=datetime(2026, 8, 10, 11, tzinfo=UTC),
        revision_id=explicit_revision,
    )
    repeated = await collect(
        captured_at=datetime(2026, 8, 10, 12, tzinfo=UTC),
        revision_id=explicit_revision,
    )

    assert legacy.revision_id.startswith("sha256:")
    assert upgraded.revision_id.startswith(f"{explicit_revision}@capture-")
    assert upgraded.evidence_hash == legacy.evidence_hash
    assert repeated.revision_id == upgraded.revision_id
    assert repeated.knowledge_time == upgraded.knowledge_time
